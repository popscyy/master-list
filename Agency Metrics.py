from io import BytesIO
import re
from pathlib import Path
from datetime import date, datetime

import pandas as pd
import polars as pl
import streamlit as st
import msoffcrypto
from openpyxl import load_workbook


DEFAULT_LOCAL_OPTIONS_DIR = "options"
DEFAULT_ENDORSEMENT_FILE = "MAYA ENDORSEMENT 04082026.xlsx"
DEFAULT_WORKBOOK_PASSWORD = "Maya@2026"
DEFAULT_SERVER_MASTERFILE_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\ENDO\MASTERFILE"


class LocalInputFile:
    def __init__(self, file_path: Path):
        self.path = file_path
        self.name = file_path.name

    def getvalue(self) -> bytes:
        return self.path.read_bytes()


def resolve_server_endorsement_file(server_input: str) -> tuple[Path | None, str | None]:
    master_dir = Path(DEFAULT_SERVER_MASTERFILE_DIR)
    if not master_dir.exists() or not master_dir.is_dir():
        return None, f"Server folder is not reachable: {master_dir}"

    requested = (server_input or "").strip()
    if not requested:
        return None, "Server file mode is enabled but no file name/path was provided."

    normalized_requested = requested.replace("/", "\\")
    requested_path = Path(normalized_requested)

    # Allow a full file path directly.
    if requested_path.exists() and requested_path.is_file():
        return requested_path, None

    # Allow a path relative to MASTERFILE (e.g., MARCH 2026\\file.xlsx).
    candidate = master_dir / normalized_requested
    if candidate.exists() and candidate.is_file():
        return candidate, None

    # Fallback: search by filename anywhere under MASTERFILE.
    requested_name = requested_path.name
    if not requested_name:
        return None, "Please provide a valid server file name or relative path."

    matches = [
        p for p in master_dir.rglob("*") if p.is_file() and p.name.lower() == requested_name.lower()
    ]
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        preview = "\n".join(str(p) for p in matches[:5])
        return None, (
            f"Multiple files named '{requested_name}' were found under MASTERFILE. "
            "Please paste a relative path, for example 'MARCH 2026\\filename.xlsx'.\n"
            f"Matches:\n{preview}"
        )

    return None, f"Server file not found from input: {requested}"


def get_local_endorsement_files() -> list[LocalInputFile]:
    script_dir = Path(__file__).resolve().parent
    project_dir = script_dir.parent

    options_dir = project_dir / DEFAULT_LOCAL_OPTIONS_DIR
    scan_dirs = []
    if options_dir.exists() and options_dir.is_dir():
        scan_dirs.append(options_dir)
    scan_dirs.append(project_dir)

    all_files: list[Path] = []
    for base_dir in scan_dirs:
        all_files.extend(
            [
                p
                for p in base_dir.rglob("*")
                if p.is_file()
                and p.suffix.lower() in {".xlsx", ".xls", ".xlsb", ".csv"}
                and not p.name.startswith("~$")
                and not p.name.startswith("~")
            ]
        )

    endorsement = [
        LocalInputFile(p)
        for p in all_files
        if "endorsement" in p.name.lower()
    ]
    endorsement.sort(
        key=lambda file_obj: (
            file_obj.name.lower() != DEFAULT_ENDORSEMENT_FILE.lower(),
            file_obj.name.lower(),
        )
    )
    return endorsement


def make_unique_columns(columns: list) -> list[str]:
    seen = {}
    unique = []
    for raw_name in columns:
        base = str(raw_name).strip() if raw_name is not None else ""
        base = base if base else "COLUMN"
        count = seen.get(base, 0) + 1
        seen[base] = count
        unique.append(base if count == 1 else f"{base}_{count}")
    return unique


def standardize_column_name(name: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "_", str(name).strip().upper()).strip("_")


def pick_column(columns: list[str], candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return None


def _normalize_excel_value(value):
    if value is None:
        return None
    if isinstance(value, (bytes, bytearray)):
        try:
            return value.decode("utf-8")
        except Exception:
            return value.decode("latin-1", errors="ignore")
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    # Keep unsupported Excel cell objects as strings to avoid mixed-type crashes.
    return str(value)


def _workbook_to_polars(workbook: dict) -> pl.DataFrame:
    frames = []
    for sheet_name, pdf in workbook.items():
        if pdf is None or pdf.empty:
            continue

        # Normalize strange cell types before Polars conversion.
        for col in pdf.columns:
            if pdf[col].dtype == "object":
                pdf[col] = pdf[col].map(_normalize_excel_value)

        pdf.columns = make_unique_columns(pdf.columns)
        pdf["_SOURCE_SHEET"] = str(sheet_name)
        frames.append(pl.from_pandas(pdf))

    if not frames:
        raise ValueError("Endorsement workbook has no readable rows.")

    return pl.concat(frames, how="diagonal_relaxed")


def _worksheet_to_polars(ws) -> pl.DataFrame | None:
    preview_rows = list(ws.iter_rows(min_row=1, max_row=40, values_only=True))
    if not preview_rows:
        return None

    def row_score(row) -> int:
        tokens = {
            str(cell).strip().upper()
            for cell in row
            if cell is not None and str(cell).strip() != ""
        }
        markers = {
            "PLACEMENT",
            "ACCOUNT NUMBER",
            "RECEIVED DATE",
            "SUB CAMPAIGN",
            "ENDO DATE",
            "PAYMENT AMOUNT",
        }
        return len(tokens & markers)

    header_idx = 0
    best_score = -1
    for idx, row in enumerate(preview_rows):
        score = row_score(row)
        if score > best_score:
            best_score = score
            header_idx = idx

    header = preview_rows[header_idx]
    if header is None:
        return None

    header = make_unique_columns(list(header))
    width = len(header)

    data_rows = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        normalized = [_normalize_excel_value(value) for value in list(row[:width])]
        if len(normalized) < width:
            normalized.extend([None] * (width - len(normalized)))
        if any(value is not None and str(value).strip() != "" for value in normalized):
            data_rows.append(normalized)

    if not data_rows:
        return None

    data = {name: [row[idx] for row in data_rows] for idx, name in enumerate(header)}
    frame = pl.DataFrame(data, strict=False)
    return frame.with_columns(pl.lit(str(ws.title)).alias("_SOURCE_SHEET"))


def _read_decrypted_workbook(file_bytes: bytes, workbook_password: str) -> pl.DataFrame:
    decrypted = BytesIO()
    office = msoffcrypto.OfficeFile(BytesIO(file_bytes))
    office.load_key(password=workbook_password)
    office.decrypt(decrypted)
    decrypted.seek(0)

    workbook = load_workbook(decrypted, read_only=True, data_only=True)
    frames = []
    for ws in workbook.worksheets:
        frame = _worksheet_to_polars(ws)
        if frame is not None:
            frames.append(frame)

    if not frames:
        raise ValueError("Decrypted workbook has no readable sheets.")

    return pl.concat(frames, how="diagonal_relaxed")


def read_endorsement_file(file_obj, workbook_password: str | None = None):
    ext = Path(file_obj.name).suffix.lower()
    file_bytes = file_obj.getvalue()

    if ext == ".csv":
        pdf = pd.read_csv(BytesIO(file_bytes))
        pdf.columns = make_unique_columns(pdf.columns)
        pdf["_SOURCE_SHEET"] = "CSV"
        return pl.from_pandas(pdf)

    effective_password = workbook_password or DEFAULT_WORKBOOK_PASSWORD

    if effective_password:
        try:
            return _read_decrypted_workbook(file_bytes, effective_password)
        except Exception as exc:
            raise ValueError(f"Could not read workbook after decrypting with the provided password: {exc}")

    raise ValueError("Workbook password is required for this endorsement file.")


def to_text_expr(df: pl.DataFrame, column_name: str | None) -> pl.Expr:
    if column_name is None or column_name not in df.columns:
        return pl.lit("")
    return pl.col(column_name).cast(pl.Utf8).fill_null("")


def to_date_expr(df: pl.DataFrame, column_name: str | None) -> pl.Expr:
    if column_name is None or column_name not in df.columns:
        return pl.lit(None, dtype=pl.Date)

    txt = pl.col(column_name).cast(pl.Utf8).str.strip_chars().str.replace_all(r"\s+", " ")
    return pl.coalesce(
        [
            pl.col(column_name).cast(pl.Date, strict=False),
            pl.col(column_name).cast(pl.Datetime, strict=False).dt.date(),
            txt.str.strptime(pl.Date, "%m/%d/%Y", strict=False),
            txt.str.strptime(pl.Date, "%Y-%m-%d", strict=False),
            txt.str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S", strict=False).dt.date(),
            txt.str.strptime(pl.Datetime, "%m/%d/%Y %H:%M:%S", strict=False).dt.date(),
            txt.str.strptime(pl.Date, "%d/%m/%Y", strict=False),
        ]
    )


def prepare_endorsement(endo_df: pl.DataFrame) -> pl.DataFrame:
    if endo_df.is_empty():
        raise ValueError("Endorsement data is empty.")

    std_names = [standardize_column_name(col) for col in endo_df.columns]
    rename_map = {}
    seen = {}
    for old_col, std_col in zip(endo_df.columns, std_names):
        count = seen.get(std_col, 0) + 1
        seen[std_col] = count
        rename_map[old_col] = std_col if count == 1 else f"{std_col}_{count}"

    df = endo_df.rename(rename_map)

    account_col = pick_column(df.columns, ["ACCOUNT_NO", "ACCOUNT_NUM", "ACCOUNT", "ACCOUNT_NUMBER"])
    agency_col = pick_column(df.columns, ["AGENCY", "AGENCY_NAME", "ENDORSEMENT_AGENCY", "PLACEMENT"])
    received_date_col = pick_column(df.columns, ["RECEIVED_DATE", "DATE_OF_ASSIGNMENT", "AS_OF"])
    pulled_out_date_col = pick_column(df.columns, ["PULLED_OUT_DATE", "PULLED OUT DATE", "PULLED_OUT", "PULLED DATE"])
    source_sheet_col = pick_column(df.columns, ["SOURCE_SHEET", "_SOURCE_SHEET"])
    ob_col = pick_column(df.columns, ["OB", "OUTSTANDING_BALANCE", "BALANCE", "OSB"])
    campaign_col = pick_column(df.columns, ["CAMPAIGN", "CAMPAIGN_NAME"])
    sub_campaign_col = pick_column(df.columns, ["SUB_CAMPAIGN", "SUBCAMPAIGN", "CAMPAIGN"])
    payment_date_col = pick_column(df.columns, ["PAYMENT_DATE", "PAYMENT DATE", "DATE"])
    endo_date_col = pick_column(df.columns, ["ENDO_DATE", "END_DATE", "PAYMENT_DATE", "DATE"])
    payment_amount_col = pick_column(
        df.columns,
        ["PAYMENT_AMOUNT", "AMOUNT", "COLLECTED_AMOUNT", "COLLECTED", "PAID_AMOUNT"],
    )

    if account_col is None:
        raise ValueError("Missing account column. Expected ACCOUNT/ACCOUNT_NO/ACCOUNT_NUMBER.")

    return df.with_columns(
        [
            to_text_expr(df, account_col).alias("ACCOUNT_KEY"),
            to_text_expr(df, agency_col).alias("AGENCY_STD"),
            to_text_expr(df, source_sheet_col).str.to_uppercase().str.strip_chars().alias("SOURCE_SHEET_STD"),
            to_date_expr(df, received_date_col).alias("RECEIVED_DATE_STD"),
            to_date_expr(df, pulled_out_date_col).alias("PULLED_OUT_DATE_STD"),
            (pl.col(ob_col).cast(pl.Float64, strict=False).fill_null(0.0) if ob_col else pl.lit(0.0)).alias("OB_STD"),
            to_text_expr(df, campaign_col).str.to_uppercase().str.strip_chars().alias("CAMPAIGN_STD"),
            to_text_expr(df, sub_campaign_col).str.to_uppercase().str.strip_chars().alias("SUB_CAMPAIGN_STD"),
            to_date_expr(df, payment_date_col).alias("PAYMENT_DATE_STD"),
            to_date_expr(df, endo_date_col).alias("ENDO_DATE_STD"),
            (pl.col(payment_amount_col).cast(pl.Float64, strict=False).fill_null(0.0) if payment_amount_col else pl.lit(0.0)).alias("PAYMENT_AMOUNT_STD"),
        ]
    )


def calculate_endorsement_metrics(
    df: pl.DataFrame,
    source_sheet: str,
    month: int,
    year: int,
    use_endorsement_rules: bool,
) -> pl.DataFrame:
    source_value = source_sheet.strip().upper()
    if use_endorsement_rules:
        filtered = df.filter(
            (pl.col("SOURCE_SHEET_STD") == source_value)
            & (pl.col("RECEIVED_DATE_STD").is_not_null())
            & (pl.col("RECEIVED_DATE_STD").dt.month() == month)
            & (pl.col("RECEIVED_DATE_STD").dt.year() == year)
        )
        metric_name = "#_OF_ENDORSED_ACCOUNTS_HANDLED"
        bucket_label = f"{source_value} / RULES"
    else:
        filtered = df
        metric_name = "#_OF_ENDORSED_ACCOUNTS_HANDLED_DEFAULT"
        bucket_label = "DEFAULT FORMAT"

    unique_accounts = filtered.select(pl.col("ACCOUNT_KEY").n_unique()).item() if filtered.height > 0 else 0

    return pl.DataFrame(
        [
            {
                "METRIC": metric_name,
                "SOURCE_SHEET": bucket_label,
                "MONTH": month,
                "YEAR": year,
                "VALUE": int(unique_accounts),
            }
        ]
    )


def calculate_agency_input_table_121_150(
    prepared_df: pl.DataFrame,
    month: int,
    year: int,
) -> pl.DataFrame:
    target_placement = "MAYA CREDIT 121 - 150 DPD"
    normalized_target_placement = re.sub(r"\s+", " ", target_placement.strip().upper())
    placement_expr = pl.col("AGENCY_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()
    sheet_expr = pl.col("SOURCE_SHEET_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.strip_chars()
    sub_campaign_expr = pl.col("SUB_CAMPAIGN_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()

    # 1) Endorsed handled: ACTIVE sheet rows filtered by RECEIVED DATE + POUT sheet rows filtered by RECEIVED DATE, then combine and dedupe by account.
    active_filtered = prepared_df.filter(
        (sheet_expr == "ACTIVE")
        & (placement_expr == normalized_target_placement)
        & pl.col("RECEIVED_DATE_STD").is_not_null()
        & (pl.col("RECEIVED_DATE_STD").dt.year() == year)
        & (pl.col("RECEIVED_DATE_STD").dt.month() == month)
    )
    pout_filtered = prepared_df.filter(
        (sheet_expr == "POUT")
        & (placement_expr == normalized_target_placement)
        & pl.col("RECEIVED_DATE_STD").is_not_null()
        & (pl.col("RECEIVED_DATE_STD").dt.year() == year)
        & (pl.col("RECEIVED_DATE_STD").dt.month() == month)
    )
    endorsed_filtered = pl.concat([active_filtered, pout_filtered], how="diagonal_relaxed")
    endorsed_count = (
        endorsed_filtered.filter(pl.col("ACCOUNT_KEY").str.strip_chars() != "")
        .select(pl.col("ACCOUNT_KEY").n_unique())
        .item()
        if endorsed_filtered.height > 0
        else 0
    )

    # 2) Pulled out: POUT sheet rows filtered by placement/date, then count rows.
    pulled_out_filtered = pout_filtered
    pulled_out_count = pulled_out_filtered.height

    # 3) OSB Endorsed: dedupe combined ACTIVE + POUT by account, then sum OB.
    osb_filtered = endorsed_filtered.unique(subset=["ACCOUNT_KEY"], keep="first")
    osb_endorsed = osb_filtered.select(pl.col("OB_STD").sum()).item() if osb_filtered.height > 0 else 0.0

    # 4) Collected amount: PAYMENTS + SELECTIVES for the 121 month/year, both filtered by ENDO DATE.
    payments_121 = prepared_df.filter(
        (sheet_expr == "PAYMENTS")
        & (sub_campaign_expr == normalized_target_placement)
        & pl.col("ENDO_DATE_STD").is_not_null()
        & (pl.col("ENDO_DATE_STD").dt.year() == year)
        & (pl.col("ENDO_DATE_STD").dt.month() == month)
    )
    selectives_121 = prepared_df.filter(
        (sheet_expr == "SELECTIVES")
        & (sub_campaign_expr == normalized_target_placement)
        & pl.col("ENDO_DATE_STD").is_not_null()
        & (pl.col("ENDO_DATE_STD").dt.year() == year)
        & (pl.col("ENDO_DATE_STD").dt.month() == month)
    )

    payments_sum = payments_121.select(pl.col("PAYMENT_AMOUNT_STD").sum()).item() if payments_121.height > 0 else 0.0
    selectives_sum = selectives_121.select(pl.col("PAYMENT_AMOUNT_STD").sum()).item() if selectives_121.height > 0 else 0.0
    collected_amount = float((payments_sum or 0.0) + (selectives_sum or 0.0))

    st.caption(
        "Collected breakdown - PAYMENTS 121-150 "
        f"({month_name(month)}): {payments_sum or 0.0:,.2f} | SELECTIVES 121-150 "
        f"({month_name(month)}): {selectives_sum or 0.0:,.2f}"
    )

    return pl.DataFrame(
        [
            {
                "Bucket": "121-150 DPD",
                "Agency": "SP MADRID",
                "# of Endorsed Accounts Handled": int(endorsed_count),
                "# of Accounts Pulled Out": int(pulled_out_count),
                "OSB Endorsed (₱)": float(osb_endorsed or 0.0),
                "Collected (₱)": collected_amount,
            }
        ]
    )


def export_summary_excel(summary_df: pl.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_pandas().to_excel(writer, index=False, sheet_name="Summary")
    output.seek(0)
    return output.getvalue()

def month_name(month: int) -> str:
    return date(2000, int(month), 1).strftime("%B")


def previous_month_pair(year: int, month: int) -> tuple[int, int]:
    if month == 1:
        return 12, year - 1
    return month - 1, year


def calculate_agency_input_table_181_above(
    prepared_df: pl.DataFrame,
    collected_month: int,
    collected_year: int,
) -> pl.DataFrame:
    target_placement = "MAYA CREDIT 181 DPD & UP"
    normalized_target_placement = re.sub(r"\s+", " ", target_placement.strip().upper())
    placement_expr = pl.col("AGENCY_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()
    sheet_expr = pl.col("SOURCE_SHEET_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.strip_chars()
    sub_campaign_expr = pl.col("SUB_CAMPAIGN_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()
    campaign_expr = pl.col("CAMPAIGN_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()

    # 1) Endorsed handled: ACTIVE sheet rows for 181+.
    active_filtered = prepared_df.filter(
        (sheet_expr == "ACTIVE")
        & (placement_expr == normalized_target_placement)
    )

    # 2) Pulled out: POUT sheet rows for 181+ filtered by pulled out date.
    pout_endorsed_filtered = prepared_df.filter(
        (sheet_expr == "POUT")
        & (placement_expr == normalized_target_placement)
        & pl.col("PULLED_OUT_DATE_STD").is_not_null()
        & (pl.col("PULLED_OUT_DATE_STD").dt.year() == collected_year)
        & (pl.col("PULLED_OUT_DATE_STD").dt.month() == collected_month)
    )
    endorsed_filtered = pl.concat([active_filtered, pout_endorsed_filtered], how="diagonal_relaxed")
    endorsed_count = (
        endorsed_filtered.filter(pl.col("ACCOUNT_KEY").str.strip_chars() != "")
        .select(pl.col("ACCOUNT_KEY").n_unique())
        .item()
        if endorsed_filtered.height > 0
        else 0
    )

    pulled_out_count = pout_endorsed_filtered.height

    # 3) OSB Endorsed: dedupe combined ACTIVE + POUT by account, then sum OB.
    osb_filtered = endorsed_filtered.unique(subset=["ACCOUNT_KEY"], keep="first")
    osb_endorsed = osb_filtered.select(pl.col("OB_STD").sum()).item() if osb_filtered.height > 0 else 0.0

    # 4) Collected amount: PAYMENTS for 181+ in the selected month/year.
    payments_181 = prepared_df.filter(
        sheet_expr.is_in(["PAYMENT", "PAYMENTS"])
        & (sub_campaign_expr == normalized_target_placement)
        & (campaign_expr == "MAYA CREDIT")
        & pl.col("PAYMENT_DATE_STD").is_not_null()
        & (pl.col("PAYMENT_DATE_STD").dt.month() == collected_month)
        & (pl.col("PAYMENT_DATE_STD").dt.year() == collected_year)
    )

    payments_sum = payments_181.select(pl.col("PAYMENT_AMOUNT_STD").sum()).item() if payments_181.height > 0 else 0.0
    collected_amount = float(payments_sum or 0.0)

    st.caption(
        f"Collected breakdown - PAYMENTS 181 {month_name(collected_month)} only: "
        f"{payments_sum or 0.0:,.2f}"
    )

    return pl.DataFrame(
        [
            {
                "Bucket": "181+ DPD",
                "Agency": "SP MADRID",
                "# of Endorsed Accounts Handled": int(endorsed_count),
                "# of Accounts Pulled Out": int(pulled_out_count),
                "OSB Endorsed (₱)": float(osb_endorsed or 0.0),
                "Collected (₱)": collected_amount,
            }
        ]
    )


def run_ui():
    st.header("Agency Metrics")
    st.caption("Endorsement-only processing")

    with st.expander("Options", expanded=True):
        use_server_file = st.checkbox("Use server file by name", value=True)
        server_file_name = ""
        if use_server_file:
            st.caption(f"Server source: {DEFAULT_SERVER_MASTERFILE_DIR}")
            server_file_name = st.text_input(
                "Server file name",
                value=DEFAULT_ENDORSEMENT_FILE,
                help="Paste filename, relative path (e.g. MARCH 2026\\file.xlsx), or full UNC file path.",
            )
        use_default_calculation = st.checkbox(
            "Use default calculation",
            value=False,
            help="Checked uses the current fixed-month default calculation. Unchecked lets you pick a month/year and uses previous-month flow for 181.",
        )
        if not use_default_calculation:
            month = st.selectbox(
                "Upload month",
                options=list(range(1, 13)),
                index=3,
                format_func=month_name,
            )
            year = st.number_input("Upload year", min_value=2000, max_value=2100, value=2026, step=1)
            current_month = int(month)
            current_year = int(year)
            previous_month, previous_year = previous_month_pair(current_year, current_month)
        else:
            current_month = 3
            current_year = 2026
            previous_month, previous_year = previous_month_pair(current_year, current_month)
        workbook_password = st.text_input(
            "Workbook password",
            value=DEFAULT_WORKBOOK_PASSWORD,
            type="password",
        )

    st.write("Upload endorsement files (optional if local folder mode is enabled).")
    uploaded_files = st.file_uploader(
        "Upload Endorsement File(s)",
        type=["xlsx", "xls", "xlsb", "csv"],
        accept_multiple_files=True,
    )

    if st.button("Submit", type="secondary", use_container_width=True):
        progress = st.progress(0)
        status = st.empty()

        def set_progress(percent: int, message: str):
            progress.progress(percent)
            status.caption(f"{percent}% - {message}")

        try:
            set_progress(10, "Collecting files")
            selected_file = None

            if use_server_file:
                resolved_server_file, resolve_error = resolve_server_endorsement_file(server_file_name)
                if resolve_error:
                    set_progress(100, "Stopped")
                    st.error(resolve_error)
                    return

                selected_file = LocalInputFile(resolved_server_file)
                st.caption(f"Using server file: {selected_file.path}")

            # If user uploads a file, always prioritize it over local auto-pick.
            if selected_file is None and uploaded_files:
                selected_file = uploaded_files[0]
                st.caption(f"Using uploaded file: {selected_file.name}")

            if selected_file is None:
                set_progress(100, "Stopped")
                st.error("No endorsement file selected. Upload a file or enable server file mode with a valid file name.")
                return

            set_progress(35, "Reading endorsement workbook")
            raw_df = read_endorsement_file(
                selected_file,
                workbook_password=workbook_password.strip() or None,
            )

            set_progress(65, "Preparing endorsement columns")
            prepared_df = prepare_endorsement(raw_df)


            set_progress(85, "Calculating metrics")
            st.subheader("Excel Preview")
            if use_default_calculation:
                st.caption(f"Previewing {selected_file.name} using the current calculation.")
            else:
                st.caption(
                    f"Previewing {selected_file.name} with 121 set to {month_name(int(previous_month))} and 181 set to {month_name(int(current_month))}."
                )
            st.dataframe(prepared_df.to_pandas().head(50), use_container_width=True)

            st.subheader("Agency Input Calculation Table")
            st.caption(
                f"Month flow: 121 = {month_name(int(previous_month))}, 181 = {month_name(int(current_month))}."
            )
            calc_table_121_150 = calculate_agency_input_table_121_150(
                prepared_df,
                month=int(previous_month),
                year=int(previous_year),
            )
            calc_table_181_above = calculate_agency_input_table_181_above(
                prepared_df,
                collected_month=int(current_month),
                collected_year=int(current_year),
            )
            combined_table = pl.concat([calc_table_121_150, calc_table_181_above], how="diagonal_relaxed")
            calc_pd = combined_table.to_pandas()
            st.dataframe(
                calc_pd.style.format(
                    {
                        "# of Endorsed Accounts Handled": "{:,.0f}",
                        "# of Accounts Pulled Out": "{:,.0f}",
                        "OSB Endorsed (₱)": "{:,.2f}",
                        "Collected (₱)": "{:,.2f}",
                    }
                ),
                use_container_width=True,
            )

            set_progress(100, "Completed")
            st.success("Endorsement metrics generated successfully.")
        except Exception as exc:
            set_progress(100, "Failed")
            st.error(f"Failed to process endorsement file: {exc}")


run_ui()

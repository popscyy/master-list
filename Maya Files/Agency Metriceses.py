import os
import re
import tempfile
import subprocess
import streamlit as st
import pandas as pd
import polars as pl
import msoffcrypto
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, Tuple
from dotenv import load_dotenv
from openpyxl import load_workbook


def _ensure_network_access(path: str):
    """Check if network path is accessible."""
    try:
        if Path(path).exists():
            return True
    except Exception:
        pass
    return False
    
load_dotenv()

st.header("Agency Metrics")

# ── Constants ────────────────────────────────────────────────────────────────
STATUS_COLUMN = "status"
CYCLE_COLUMN  = "cycle"

METRIC_LABELS = {
    "dials":            "Dials",
    "manual_dials":     "Manual Dials",
    "rpc_under_nego":   "RPC Under Nego",
    "rpc_niop":         "RPC NIOP",
    "third_party":      "3rd Party Contacted",
    "dispute":          "Dispute",
    "email_sent":       "Email Sent",
    "email_responsive": "Email Responsive",
    "vb_count":         "VB Count",
    "vb_connected":     "VB Connected",
}

# ── Masterfile / Endorsement Logic ──────────────────────────────────────────
MF_PATH    = os.getenv("MASTERFILE_PATH", "")
MF_WB_PASS = os.getenv("MASTERFILE_WORKBOOK_PASSWORD", "Maya@2026")
DEFAULT_SERVER_MASTERFILE_DIR = os.getenv(
    "DEFAULT_SERVER_MASTERFILE_DIR",
    r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\ENDO\MASTERFILE",
)
DEFAULT_SERVER_ENDORSEMENT_FILE = f"MAYA ENDORSEMENT {date.today():%m%d%Y}.xlsx"
MASTERFILE_SHEETS = ("ACTIVE", "POUT", "PAYMENTS", "SELECTIVES")


def resolve_server_endorsement_file(server_input: str) -> tuple[Path | None, str | None]:
    master_dir = Path(DEFAULT_SERVER_MASTERFILE_DIR)
    if not master_dir.exists() or not master_dir.is_dir():
        return None, f"Server folder is not reachable: {master_dir}"

    requested = (server_input or "").strip()
    if not requested:
        return None, "Server file mode is enabled but no file name/path was provided."

    normalized_requested = requested.replace("/", "\\")
    requested_path = Path(normalized_requested)

    # Allow a full file path directly.
    if requested_path.exists() and requested_path.is_file():
        return requested_path, None

    # Allow a path relative to MASTERFILE (e.g., MARCH 2026\\file.xlsx).
    candidate = master_dir / normalized_requested
    if candidate.exists() and candidate.is_file():
        return candidate, None

    # Fallback: search by filename anywhere under MASTERFILE.
    requested_name = requested_path.name
    if not requested_name:
        return None, "Please provide a valid server file name or relative path."

    matches = [
        p for p in master_dir.rglob("*") if p.is_file() and p.name.lower() == requested_name.lower()
    ]
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        preview = "\n".join(str(p) for p in matches[:5])
        return None, (
            f"Multiple files named '{requested_name}' were found under MASTERFILE. "
            "Please paste a relative path, for example 'MARCH 2026\\file.xlsx'.\n"
            f"Matches:\n{preview}"
        )

    return None, f"Server file not found from input: {requested}"


def get_latest_endorsement_file() -> Path | None:
    """Try to find the latest endorsement file in the configured path (if accessible)."""
    if not MF_PATH:
        return None
    try:
        folder = Path(MF_PATH)
        if not folder.exists():
            return None
        # Search recursively for endorsement files (handles monthly subfolders)
        files = list(folder.rglob("*endorsement*.xlsx")) + \
                list(folder.rglob("*endorsement*.xls")) + \
                list(folder.rglob("*endorsement*.xlsb"))
        # Filter out temp files
        files = [f for f in files if not f.name.startswith("~")]
        return max(files, key=lambda p: p.stat().st_mtime) if files else None
    except Exception:
        return None


def standardize_column_name(name: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "_", str(name).strip().upper()).strip("_")


def pick_column(columns: list[str], candidates: list[str]) -> str | None:
    for c in candidates:
        if c in columns:
            return c
    return None


def _file_to_bytes(file_obj) -> bytes:
    if isinstance(file_obj, Path):
        return file_obj.read_bytes()
    if hasattr(file_obj, "getvalue"):
        return file_obj.getvalue()
    if hasattr(file_obj, "read"):
        return file_obj.read()
    raise TypeError(f"Unsupported file object: {type(file_obj)!r}")


def _decrypt_excel_bytes(file_bytes: bytes, wb_password: str) -> bytes:
    decrypted = BytesIO()
    office = msoffcrypto.OfficeFile(BytesIO(file_bytes))
    office.load_key(password=wb_password)
    office.decrypt(decrypted)
    return decrypted.getvalue()
def _canonical_masterfile_sheet_name(sheet_name: str) -> str:
    return sheet_name.upper().strip()


def _normalize_payments_dump_row(frame: pl.DataFrame) -> pl.DataFrame:
    expected = {"PAYMENT_AMOUNT_STD", "PAYMENT_DATE_STD", "SUB_CAMPAIGN_STD", "CAMPAIGN_STD"}
    cols = {standardize_column_name(c) for c in frame.columns}
    if expected & cols:
        return frame
    return pl.DataFrame()


def _read_excel_sheet(decrypted_bytes: bytes, sheet_name: str) -> pl.DataFrame | None:
    for engine in ("calamine", "openpyxl"):
        try:
            frame = pl.read_excel(
                BytesIO(decrypted_bytes),
                sheet_name=sheet_name,
                engine=engine,
            )
            if sheet_name.upper() == "PAYMENTS":
                validated = _normalize_payments_dump_row(frame)
                if validated.height == 0:
                    try:
                        repaired = pd.read_excel(
                            BytesIO(decrypted_bytes),
                            sheet_name=sheet_name,
                            engine="openpyxl",
                            header=1,
                        )
                        if repaired is not None and not repaired.empty:
                            frame = pl.from_pandas(repaired, include_index=False, rechunk=False)
                    except Exception:
                        pass
            return frame.with_columns(pl.lit(_canonical_masterfile_sheet_name(sheet_name)).alias("_SOURCE_SHEET"))
        except Exception:
            continue
    return None


@st.cache_data(show_spinner=False)
def _read_endorsement_file_cached(file_bytes: bytes, wb_password: str) -> pl.DataFrame:
    decrypted_bytes = _decrypt_excel_bytes(file_bytes, wb_password)
    workbook = load_workbook(BytesIO(decrypted_bytes), read_only=True, data_only=True)
    available_sheets = {name.upper(): name for name in workbook.sheetnames}
    target_sheets = [available_sheets[name] for name in MASTERFILE_SHEETS if name in available_sheets]
    workbook.close()

    if not target_sheets:
        raise ValueError("No readable sheets found in workbook.")

    frames = []
    for sheet_name in MASTERFILE_SHEETS:
        if sheet_name in target_sheets:
            table = _read_excel_sheet(decrypted_bytes, sheet_name)
            if table is not None and table.height > 0:
                frames.append(table)

    if not frames:
        raise ValueError("No readable sheets found in workbook.")
    return pl.concat(frames, how="diagonal_relaxed")


def read_endorsement_file(file_obj, wb_password: str) -> pl.DataFrame:
    return _read_endorsement_file_cached(_file_to_bytes(file_obj), wb_password)


def to_text_expr(df: pl.DataFrame, col: str | None) -> pl.Expr:
    if col is None or col not in df.columns:
        return pl.lit("")
    return pl.col(col).cast(pl.Utf8).fill_null("")


def to_date_expr(df: pl.DataFrame, col: str | None) -> pl.Expr:
    if col is None or col not in df.columns:
        return pl.lit(None, dtype=pl.Date)
    txt = pl.col(col).cast(pl.Utf8).str.strip_chars().str.replace_all(r"\s+", " ")
    return pl.coalesce([
        pl.col(col).cast(pl.Date, strict=False),
        pl.col(col).cast(pl.Datetime, strict=False).dt.date(),
        txt.str.strptime(pl.Date, "%m/%d/%Y",          strict=False),
        txt.str.strptime(pl.Date, "%Y-%m-%d",          strict=False),
        txt.str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S", strict=False).dt.date(),
        txt.str.strptime(pl.Datetime, "%m/%d/%Y %H:%M:%S", strict=False).dt.date(),
        txt.str.strptime(pl.Date, "%d/%m/%Y",          strict=False),
    ])


def prepare_endorsement(endo_df: pl.DataFrame) -> pl.DataFrame:
    std_names  = [standardize_column_name(c) for c in endo_df.columns]
    seen, rmap = {}, {}
    for old, std in zip(endo_df.columns, std_names):
        count      = seen.get(std, 0) + 1
        seen[std]  = count
        rmap[old]  = std if count == 1 else f"{std}_{count}"
    df = endo_df.rename(rmap)

    account_col      = pick_column(df.columns, ["ACCOUNT_NO", "ACCOUNT_NUM", "ACCOUNT", "ACCOUNT_NUMBER"])
    agency_col       = pick_column(df.columns, ["AGENCY", "AGENCY_NAME", "ENDORSEMENT_AGENCY", "PLACEMENT"])
    received_col     = pick_column(df.columns, ["RECEIVED_DATE", "DATE_OF_ASSIGNMENT", "AS_OF"])
    pulled_out_col   = pick_column(df.columns, ["PULLED_OUT_DATE", "PULLED OUT DATE", "PULLED_OUT", "PULLED_DATE"])
    source_sheet_col = pick_column(df.columns, ["SOURCE_SHEET", "_SOURCE_SHEET"])
    ob_col           = pick_column(df.columns, ["OB", "OUTSTANDING_BALANCE", "BALANCE", "OSB"])
    campaign_col     = pick_column(df.columns, ["CAMPAIGN", "CAMPAIGN_NAME"])
    sub_campaign_col = pick_column(df.columns, ["SUB_CAMPAIGN", "SUBCAMPAIGN", "CAMPAIGN"])
    payment_date_col = pick_column(df.columns, ["PAYMENT_DATE", "PAYMENT DATE", "DATE"])
    endo_date_col    = pick_column(df.columns, ["ENDO_DATE", "END_DATE", "PAYMENT_DATE", "DATE"])
    payment_amt_col  = pick_column(df.columns, ["PAYMENT_AMOUNT", "AMOUNT", "COLLECTED_AMOUNT", "COLLECTED", "PAID_AMOUNT"])

    if account_col is None:
        raise ValueError("Missing account column.")

    return df.with_columns([
        to_text_expr(df, account_col).str.strip_chars().alias("ACCOUNT_KEY"),
        to_text_expr(df, agency_col).str.to_uppercase().str.strip_chars().alias("AGENCY_STD"),
        to_text_expr(df, source_sheet_col).str.to_uppercase().str.strip_chars().alias("SOURCE_SHEET_STD"),
        to_date_expr(df, received_col).alias("RECEIVED_DATE_STD"),
        to_date_expr(df, pulled_out_col).alias("PULLED_OUT_DATE_STD"),
        (pl.col(ob_col).cast(pl.Float64, strict=False).fill_null(0.0) if ob_col else pl.lit(0.0)).alias("OB_STD"),
        to_text_expr(df, campaign_col).str.to_uppercase().str.strip_chars().alias("CAMPAIGN_STD"),
        to_text_expr(df, sub_campaign_col).str.to_uppercase().str.strip_chars().alias("SUB_CAMPAIGN_STD"),
        to_date_expr(df, payment_date_col).alias("PAYMENT_DATE_STD"),
        to_date_expr(df, endo_date_col).alias("ENDO_DATE_STD"),
        (pl.col(payment_amt_col).cast(pl.Float64, strict=False).fill_null(0.0) if payment_amt_col else pl.lit(0.0)).alias("PAYMENT_AMOUNT_STD"),
    ])


def month_name(month: int) -> str:
    return date(2000, int(month), 1).strftime("%B")


def previous_month_pair(year: int, month: int) -> tuple[int, int]:
    if month == 1:
        return 12, year - 1
    return month - 1, year


def extract_month_year_from_filename(file_name: str) -> tuple[int, int]:
    match = re.search(r"(\d{8})", Path(file_name).name)
    if not match:
        raise ValueError(
            f"Could not detect MMDDYYYY in filename: {file_name}. "
            "Rename the file so it includes an 8-digit date, for example 'MAYA ENDORSEMENT 04082026.xlsx'."
        )

    parsed = datetime.strptime(match.group(1), "%m%d%Y")
    return parsed.month, parsed.year


def calc_121_150(df: pl.DataFrame, month: int, year: int) -> pl.DataFrame:
    target_placement = "MAYA CREDIT 121 - 150 DPD"
    normalized_target_placement = re.sub(r"\s+", " ", target_placement.strip().upper())
    placement_expr = pl.col("AGENCY_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()
    sheet_expr = pl.col("SOURCE_SHEET_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.strip_chars()
    sub_campaign_expr = pl.col("SUB_CAMPAIGN_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()

    active_filtered = df.filter(
        (sheet_expr == "ACTIVE")
        & (placement_expr == normalized_target_placement)
        & pl.col("RECEIVED_DATE_STD").is_not_null()
        & (pl.col("RECEIVED_DATE_STD").dt.year() == year)
        & (pl.col("RECEIVED_DATE_STD").dt.month() == month)
    )
    pout_filtered = df.filter(
        (sheet_expr == "POUT")
        & (placement_expr == normalized_target_placement)
        & pl.col("RECEIVED_DATE_STD").is_not_null()
        & (pl.col("RECEIVED_DATE_STD").dt.year() == year)
        & (pl.col("RECEIVED_DATE_STD").dt.month() == month)
    )
    endorsed_filtered = pl.concat([active_filtered, pout_filtered], how="diagonal_relaxed")

    endorsed_count = (
        endorsed_filtered.filter(pl.col("ACCOUNT_KEY").str.strip_chars() != "")
        .select(pl.col("ACCOUNT_KEY").n_unique())
        .item()
        if endorsed_filtered.height > 0
        else 0
    )

    pulled_out_count = pout_filtered.height
    osb_filtered = endorsed_filtered.unique(subset=["ACCOUNT_KEY"], keep="first")
    osb_endorsed = osb_filtered.select(pl.col("OB_STD").sum()).item() if osb_filtered.height > 0 else 0.0

    payments_121 = df.filter(
        (sheet_expr == "PAYMENTS")
        & (sub_campaign_expr == normalized_target_placement)
        & pl.col("ENDO_DATE_STD").is_not_null()
        & (pl.col("ENDO_DATE_STD").dt.year() == year)
        & (pl.col("ENDO_DATE_STD").dt.month() == month)
    )
    selectives_121 = df.filter(
        (sheet_expr == "SELECTIVES")
        & (sub_campaign_expr == normalized_target_placement)
        & pl.col("ENDO_DATE_STD").is_not_null()
        & (pl.col("ENDO_DATE_STD").dt.year() == year)
        & (pl.col("ENDO_DATE_STD").dt.month() == month)
    )

    payments_sum = payments_121.select(pl.col("PAYMENT_AMOUNT_STD").sum()).item() if payments_121.height > 0 else 0.0
    selectives_sum = selectives_121.select(pl.col("PAYMENT_AMOUNT_STD").sum()).item() if selectives_121.height > 0 else 0.0
    collected_amount = float((payments_sum or 0.0) + (selectives_sum or 0.0))

    st.caption(
        "Collected breakdown - PAYMENTS 121-150 "
        f"({month_name(month)}): {payments_sum or 0.0:,.2f} | SELECTIVES 121-150 "
        f"({month_name(month)}): {selectives_sum or 0.0:,.2f}"
    )

    return pl.DataFrame(
        [
            {
                "Bucket": "121-150 DPD",
                "Agency": "SP MADRID",
                "# of Endorsed Accounts Handled": int(endorsed_count),
                "# of Accounts Pulled Out": int(pulled_out_count),
                "OSB Endorsed (₱)": float(osb_endorsed or 0.0),
                "Collected (₱)": collected_amount,
            }
        ]
    )


def calc_181_above(df: pl.DataFrame, collected_month: int, collected_year: int) -> pl.DataFrame:
    target_placement = "MAYA CREDIT 181 DPD & UP"
    normalized_target_placement = re.sub(r"\s+", " ", target_placement.strip().upper())
    placement_expr = pl.col("AGENCY_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()
    sheet_expr = pl.col("SOURCE_SHEET_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.strip_chars()
    sub_campaign_expr = pl.col("SUB_CAMPAIGN_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()
    campaign_expr = pl.col("CAMPAIGN_STD").cast(pl.Utf8).fill_null("").str.to_uppercase().str.replace_all(r"\s+", " ").str.strip_chars()
    is_181_sub_campaign = (
        sub_campaign_expr.str.contains("181", literal=True)
        & sub_campaign_expr.str.contains("DPD", literal=True)
    )
    collected_date_expr = pl.coalesce([pl.col("ENDO_DATE_STD"), pl.col("PAYMENT_DATE_STD")])

    active_filtered = df.filter(
        (sheet_expr == "ACTIVE")
        & (placement_expr == normalized_target_placement)
    )

    pout_endorsed_filtered = df.filter(
        (sheet_expr == "POUT")
        & (placement_expr == normalized_target_placement)
        & pl.col("PULLED_OUT_DATE_STD").is_not_null()
        & (pl.col("PULLED_OUT_DATE_STD").dt.year() == collected_year)
        & (pl.col("PULLED_OUT_DATE_STD").dt.month() == collected_month)
    )
    endorsed_filtered = pl.concat([active_filtered, pout_endorsed_filtered], how="diagonal_relaxed")
    endorsed_count = (
        endorsed_filtered.filter(pl.col("ACCOUNT_KEY").str.strip_chars() != "")
        .select(pl.col("ACCOUNT_KEY").n_unique())
        .item()
        if endorsed_filtered.height > 0
        else 0
    )

    pulled_out_count = pout_endorsed_filtered.height

    osb_filtered = endorsed_filtered.unique(subset=["ACCOUNT_KEY"], keep="first")
    osb_endorsed = osb_filtered.select(pl.col("OB_STD").sum()).item() if osb_filtered.height > 0 else 0.0

    payments_181 = df.filter(
        sheet_expr.is_in(["PAYMENT", "PAYMENTS"])
        & is_181_sub_campaign
        & ((campaign_expr == "MAYA CREDIT") | (campaign_expr == ""))
        & collected_date_expr.is_not_null()
        & (collected_date_expr.dt.month() == collected_month)
        & (collected_date_expr.dt.year() == collected_year)
    )

    payments_sum = payments_181.select(pl.col("PAYMENT_AMOUNT_STD").sum()).item() if payments_181.height > 0 else 0.0
    collected_amount = float(payments_sum or 0.0)

    st.caption(
        f"Collected breakdown - PAYMENTS 181 {month_name(collected_month)} only: "
        f"{payments_sum or 0.0:,.2f}"
    )
    st.caption(
        f"181 matched rows ({month_name(collected_month)} {collected_year}): {payments_181.height:,}"
    )

    return pl.DataFrame(
        [
            {
                "Bucket": "181+ DPD",
                "Agency": "SP MADRID",
                "# of Endorsed Accounts Handled": int(endorsed_count),
                "# of Accounts Pulled Out": int(pulled_out_count),
                "OSB Endorsed (₱)": float(osb_endorsed or 0.0),
                "Collected (₱)": collected_amount,
            }
        ]
    )


USED_COLUMNS = {STATUS_COLUMN, CYCLE_COLUMN, "account_no.", "call_duration", "remark_type", "remark"}

DRR_CATEGORY_STATUS = {
    "UNDERNEGO": ["POSITIVE CONTACT - CALLBACK", "POSITIVE CONTACT - UNDERNEGO"],
    "NIOP": ["POSITIVE CONTACT - RPC REFUSE TO PAY", "POSITIVE CONTACT - DISPUTE"],
    "3RD PARTY CONTACTED": ["POSITIVE - 3RD PARTY CONTACTED"],
    "DISPUTE": ["POSITIVE CONTACT - DISPUTE"],
}


def _keep_drr_column(col_name: str) -> bool:
    normalized = col_name.lower().replace(" ", "_")
    return normalized in USED_COLUMNS


def _duration_seconds(series: pd.Series) -> pd.Series:
    parts = series.astype(str).str.split(":", expand=True)
    if parts.shape[1] >= 3:
        return (
            pd.to_numeric(parts[0], errors="coerce").fillna(0) * 3600
            + pd.to_numeric(parts[1], errors="coerce").fillna(0) * 60
            + pd.to_numeric(parts[2], errors="coerce").fillna(0)
        )
    return pd.Series(0, index=series.index)


def _extract_drr_category_rows(df_norm: pd.DataFrame, raw_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    result = {name: pd.DataFrame() for name in DRR_CATEGORY_STATUS}
    if STATUS_COLUMN not in df_norm.columns:
        return result

    working = df_norm.copy()
    if "account_no." in working.columns:
        acct = working["account_no."]
        numeric = pd.to_numeric(acct, errors="coerce")
        acct_normalized = numeric.where(numeric.isna(), numeric.astype("int64").astype(str))
        acct_normalized = acct_normalized.where(acct_normalized.notna(), acct.astype(str).str.strip())
        acct_normalized = acct_normalized.where(acct.notna(), None)
        working["account_normalized"] = acct_normalized
        working = working[working["account_normalized"].fillna("").str.startswith("6")]

    if "call_duration" in working.columns:
        connected_mask = _duration_seconds(working["call_duration"]) > 0
        connected = working[connected_mask]
    else:
        connected = pd.DataFrame(columns=working.columns)

    if connected.empty:
        return result

    connected_status = connected[STATUS_COLUMN].fillna("").str.upper()
    for sheet_name, statuses in DRR_CATEGORY_STATUS.items():
        matched_idx = connected[connected_status.isin(statuses)].index
        if len(matched_idx) > 0:
            result[sheet_name] = raw_df.loc[matched_idx].copy()

    return result

# ── DRR Tab Logic ──────────────────────────────────────────

def compute_metrics(df: pd.DataFrame) -> Dict:
    status = df[STATUS_COLUMN].fillna("").str.upper()

    has_account = "account_no." in df.columns
    if has_account:
        acct = df["account_no."]
        numeric = pd.to_numeric(acct, errors="coerce")
        acct_normalized = numeric.where(numeric.isna(), numeric.astype("int64").astype(str))
        acct_normalized = acct_normalized.where(acct_normalized.notna(), acct.astype(str).str.strip())
        acct_normalized = acct_normalized.where(acct.notna(), None)
        df = df.copy()
        df["account_normalized"] = acct_normalized
        df = df[df["account_normalized"].fillna("").str.startswith("6")]
        status = df[STATUS_COLUMN].fillna("").str.upper()

    if "call_duration" in df.columns:
        s     = df["call_duration"].astype(str)
        split = s.str.split(":", expand=True)
        if split.shape[1] >= 3:
            dur = (
                pd.to_numeric(split[0], errors="coerce").fillna(0) * 3600 +
                pd.to_numeric(split[1], errors="coerce").fillna(0) * 60 +
                pd.to_numeric(split[2], errors="coerce").fillna(0)
            )
        else:
            dur = pd.Series(0, index=df.index)
        connected        = df[dur > 0]
        connected_status = connected[STATUS_COLUMN].fillna("").str.upper()
    else:
        connected        = pd.DataFrame()
        connected_status = pd.Series(dtype=str)

    has_remark_type = "remark_type" in df.columns
    has_remark      = "remark" in df.columns

    # Single-pass status flag columns
    has_email = status.str.contains("EMAIL", na=False)
    has_sms   = status.str.contains("SMS",   na=False)
    has_untc  = status.str.contains("UNTC",  na=False)
    has_viber = status.str.contains("VIBER", na=False)
    exclude_filter = (~has_email) | (~has_sms) | (~has_untc) | (~has_viber)

    if has_remark_type:
        rt = df["remark_type"].fillna("").str.strip()
        manual_dials = int((exclude_filter & (rt == "Outgoing")).sum())
        dials        = int((exclude_filter & (rt != "Outgoing")).sum())
    else:
        manual_dials = 0
        dials        = int(exclude_filter.sum())

    account_lists = {}
    if has_account:
        account_lists = {
            "connected_accounts":        connected["account_normalized"].dropna().tolist(),
            "rpc_under_nego_accounts":   connected[connected_status.isin(["POSITIVE CONTACT - CALLBACK", "POSITIVE CONTACT - UNDERNEGO"])]["account_normalized"].dropna().tolist(),
            "rpc_niop_accounts":         connected[connected_status.isin(["POSITIVE CONTACT - RPC REFUSE TO PAY", "POSITIVE CONTACT - DISPUTE"])]["account_normalized"].dropna().tolist(),
            "third_party_accounts":      connected[connected_status == "POSITIVE - 3RD PARTY CONTACTED"]["account_normalized"].dropna().tolist(),
            "dispute_accounts":          connected[connected_status == "POSITIVE CONTACT - DISPUTE"]["account_normalized"].dropna().tolist(),
            "email_sent_accounts":       df[status.str.contains("SENT EMAIL",       na=False)]["account_normalized"].dropna().tolist(),
            "email_responsive_accounts": df[status.str.contains("EMAIL RESPONSIVE", na=False)]["account_normalized"].dropna().tolist(),
        }

    return {
        "dials":         dials,
        "manual_dials":  manual_dials,
        "connected":     len(connected),
        "vb_count":      int(df["remark"].fillna("").str.contains("broadcast", case=False, na=False).sum()) if has_remark else 0,
        "vb_connected":  int(status.isin(["PU", "PM"]).sum()),
        "account_lists": account_lists,
    }


def process_single_file(file, cycle_1: str, cycle_2: str) -> Tuple[str, Dict, int, int, int, Dict[str, pd.DataFrame]]:
    try:
        # calamine is a Rust-based reader — significantly faster than openpyxl
        raw = pd.read_excel(file, engine="calamine", dtype=str)
        raw_norm = raw.copy()
        raw_norm.columns = raw_norm.columns.str.lower().str.replace(" ", "_", regex=False)
        total_rows = len(raw)

        keep = [c for c in raw_norm.columns if c in USED_COLUMNS]
        df = raw_norm[keep]

        if STATUS_COLUMN not in df.columns:
            return file.name, {}, total_rows, 0, 0, {name: pd.DataFrame() for name in DRR_CATEGORY_STATUS}

        has_cycle     = CYCLE_COLUMN in df.columns
        cycle_results = {}
        category_frames = {name: [] for name in DRR_CATEGORY_STATUS}
        c1_rows = c2_rows = 0

        if has_cycle:
            cycle_col = df[CYCLE_COLUMN].astype(str).str.strip()

        if cycle_1:
            df_c1   = df[cycle_col == cycle_1] if has_cycle else df
            c1_rows = len(df_c1)
            cycle_results[cycle_1] = compute_metrics(df_c1)
            extracted = _extract_drr_category_rows(df_c1, raw)
            for sheet_name, frame in extracted.items():
                if not frame.empty:
                    category_frames[sheet_name].append(frame)

        if cycle_2:
            df_c2   = df[cycle_col == cycle_2] if has_cycle else df
            c2_rows = len(df_c2)
            cycle_results[cycle_2] = compute_metrics(df_c2)
            extracted = _extract_drr_category_rows(df_c2, raw)
            for sheet_name, frame in extracted.items():
                if not frame.empty:
                    category_frames[sheet_name].append(frame)

        if not cycle_1 and not cycle_2:
            c1_rows = total_rows
            cycle_results["All Data"] = compute_metrics(df)
            extracted = _extract_drr_category_rows(df, raw)
            for sheet_name, frame in extracted.items():
                if not frame.empty:
                    category_frames[sheet_name].append(frame)

        category_rows = {
            sheet_name: (pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=raw.columns))
            for sheet_name, frames in category_frames.items()
        }

        return file.name, cycle_results, total_rows, c1_rows, c2_rows, category_rows

    except Exception as e:
        st.error(f"Error processing {file.name}: {e}")
        return file.name, {}, 0, 0, 0, {name: pd.DataFrame() for name in DRR_CATEGORY_STATUS}


def aggregate_results(file_results: Dict) -> Dict:
    overall    = {}
    all_cycles = set()
    for d in file_results.values():
        all_cycles.update(d["cycle_status_counts"].keys())

    for cycle in all_cycles:
        totals            = defaultdict(int)
        account_aggregates = defaultdict(list)

        for d in file_results.values():
            if cycle in d["cycle_status_counts"]:
                metrics = d["cycle_status_counts"][cycle]
                for m in ["dials", "manual_dials", "connected", "vb_count", "vb_connected"]:
                    totals[m] += metrics.get(m, 0)
                for k, v in metrics.get("account_lists", {}).items():
                    account_aggregates[k].extend(v)

        totals["connected_unique"] = len(set(account_aggregates.get("connected_accounts", [])))
        totals["rpc_under_nego"]   = len(set(account_aggregates.get("rpc_under_nego_accounts", [])))
        totals["rpc_niop"]         = len(set(account_aggregates.get("rpc_niop_accounts", [])))
        totals["third_party"]      = len(set(account_aggregates.get("third_party_accounts", [])))
        totals["dispute"]          = len(set(account_aggregates.get("dispute_accounts", [])))
        totals["email_sent"]       = len(set(account_aggregates.get("email_sent_accounts", [])))
        totals["email_responsive"] = len(set(account_aggregates.get("email_responsive_accounts", [])))

        overall[cycle] = dict(totals)

    return overall


def build_excel(file_results: Dict, overall_counts: Dict, cycle_1: str, cycle_2: str) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Overall Summary
        overall_data = {"Metric": list(METRIC_LABELS.values())}
        for cycle, metrics in overall_counts.items():
            overall_data[cycle] = [metrics.get(k, 0) for k in METRIC_LABELS]
        pd.DataFrame(overall_data).to_excel(writer, sheet_name="Overall Summary", index=False)

        # Per-File Summary — one row per file+cycle
        per_file_rows = []
        for fname, d in file_results.items():
            for cycle, metrics in d["cycle_status_counts"].items():
                row = {"File": fname, "Cycle": cycle}
                for k, label in METRIC_LABELS.items():
                    row[label] = metrics.get(k, 0)
                per_file_rows.append(row)
        pd.DataFrame(per_file_rows).to_excel(writer, sheet_name="Per-File Summary", index=False)

        # File Statistics
        stats = [
            {
                "File":            fname,
                "Total Rows":      d["total_rows"],
                f"{cycle_1} Rows": d["cycle_1_filtered"],
                f"{cycle_2} Rows": d["cycle_2_filtered"],
            }
            for fname, d in file_results.items()
        ]
        pd.DataFrame(stats).to_excel(writer, sheet_name="File Statistics", index=False)

        for sheet_name in DRR_CATEGORY_STATUS:
            category_frames = [
                d.get("category_rows", {}).get(sheet_name)
                for d in file_results.values()
                if not d.get("category_rows", {}).get(sheet_name, pd.DataFrame()).empty
            ]
            if category_frames:
                pd.concat(category_frames, ignore_index=True).to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)

        for ws in writer.sheets.values():
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output.seek(0)
    return output.getvalue()


# ── Session State ─────────────────────────────────────────────────────────────
for key in ["drr_results", "drr_overall", "drr_excel", "drr_cycle_1", "drr_cycle_2"]:
    if key not in st.session_state:
        st.session_state[key] = None

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_masterfile, tab_drr = st.tabs(["Masterfile", "DRR"])

with tab_masterfile:
    with st.expander("Options", expanded=True):
        use_server_file = st.checkbox("Use server file by name", value=True)
        server_file_name = ""
        if use_server_file:
            st.caption(f"Server source: {DEFAULT_SERVER_MASTERFILE_DIR}")
            server_file_name = st.text_input(
                "Server file name",
                value=DEFAULT_SERVER_ENDORSEMENT_FILE,
                help="Paste filename, relative path (e.g. MARCH 2026\\file.xlsx), or full UNC file path.",
            )
    wb_pass = st.text_input("Workbook Password", value=MF_WB_PASS, type="password")

    uploaded_mf = st.file_uploader(
        "Or Upload Endorsement File",
        type=["xlsx", "xls", "xlsb"],
        help="Upload a file directly instead of using the server file",
    )

    if st.button("Load & Calculate", use_container_width=True, key="mf_submit"):
        progress = st.progress(0, text="Starting...")
        try:
            progress.progress(10, text="Locating file...")
            selected_file = None
            if use_server_file:
                selected_file, server_err = resolve_server_endorsement_file(server_file_name)
                if server_err:
                    st.error(server_err)
                    progress.progress(100, text="Failed")
                    st.stop()
                st.caption(f"Using server file: {selected_file.name}")
            elif uploaded_mf:
                selected_file = uploaded_mf
                st.caption(f"Using file: {uploaded_mf.name}")
            else:
                st.error("Please upload a file or enable server file mode.")
                progress.progress(100, text="Failed")
                st.stop()

            current_month, current_year = extract_month_year_from_filename(selected_file.name)
            previous_month, previous_year = previous_month_pair(current_year, current_month)

            progress.progress(35, text="Reading workbook...")
            raw_df = read_endorsement_file(selected_file, wb_password=wb_pass.strip())

            progress.progress(65, text="Preparing columns...")
            prepared_df = prepare_endorsement(raw_df)

            progress.progress(85, text="Calculating metrics...")
            st.caption(
                f"Month flow: 121 = {month_name(int(previous_month))}, 181 = {month_name(int(current_month))}."
            )
            table_121 = calc_121_150(prepared_df, month=int(previous_month), year=int(previous_year))
            table_181 = calc_181_above(prepared_df, collected_month=int(current_month), collected_year=int(current_year))
            combined  = pl.concat([table_121, table_181], how="diagonal_relaxed")

            progress.progress(100, text="Done")

            st.subheader("Agency Input Calculation Table")
            st.dataframe(
                combined.to_pandas().style.format({
                    "# of Endorsed Accounts Handled": "{:,.0f}",
                    "# of Accounts Pulled Out":       "{:,.0f}",
                    "OSB Endorsed (₱)":               "{:,.2f}",
                    "Collected (₱)":                  "{:,.2f}",
                }),
                use_container_width=True,
            )

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                combined.to_pandas().to_excel(writer, index=False, sheet_name="Summary")
            output.seek(0)

            st.download_button(
                label="Download Summary (.xlsx)",
                data=output.getvalue(),
                file_name=f"masterfile_summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except Exception as e:
            progress.progress(100, text="Failed")
            st.error(f"Error: {e}")

with tab_drr:
    col1, col2 = st.columns(2)
    cycle_1_input = col1.text_input("Cycle 1", value="MC 121-150 DPD")
    cycle_2_input = col2.text_input("Cycle 2", value="MC 181DPD UP")

    uploaded_files = st.file_uploader(
        "Upload Excel File/s", type=["xlsx", "xls", "xlsm"], accept_multiple_files=True
    )

    if st.button("Process", use_container_width=True, disabled=not uploaded_files):
        file_results = {}
        total_files  = len(uploaded_files)
        with st.status("Processing files...", expanded=True) as status:
            progress = st.progress(0, text="Starting...")
            for i, f in enumerate(uploaded_files):
                progress.progress(i / total_files, text=f"Processing {f.name} ({i + 1}/{total_files})...")
                fname, cycle_counts, total, c1, c2, category_rows = process_single_file(f, cycle_1_input, cycle_2_input)
                file_results[fname] = {
                    "cycle_status_counts": cycle_counts,
                    "total_rows":          total,
                    "cycle_1_filtered":    c1,
                    "cycle_2_filtered":    c2,
                    "category_rows":       category_rows,
                }
            progress.progress(1.0, text="Finalizing...")
            status.update(label="Done", state="complete", expanded=False)

        overall = aggregate_results(file_results)
        st.session_state.drr_results  = file_results
        st.session_state.drr_overall  = overall
        st.session_state.drr_excel    = build_excel(file_results, overall, cycle_1_input, cycle_2_input)
        st.session_state.drr_cycle_1  = cycle_1_input
        st.session_state.drr_cycle_2  = cycle_2_input

    if st.session_state.drr_overall:
        st.divider()

        # Summary metrics per cycle
        for cycle, metrics in st.session_state.drr_overall.items():
            st.subheader(cycle)
            rows = [{"Metric": METRIC_LABELS[k], "Value": f"{metrics.get(k, 0):,}"} for k in METRIC_LABELS]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        # File statistics
        st.subheader("File Statistics")
        c1 = st.session_state.drr_cycle_1
        c2 = st.session_state.drr_cycle_2
        stats_rows = [
            {
                "File":            fname,
                "Total Rows":      d["total_rows"],
                f"{c1} Rows":      d["cycle_1_filtered"],
                f"{c2} Rows":      d["cycle_2_filtered"],
            }
            for fname, d in st.session_state.drr_results.items()
        ]
        st.dataframe(pd.DataFrame(stats_rows), use_container_width=True, hide_index=True)

        st.download_button(
            label="Download Summary (.xlsx)",
            data=st.session_state.drr_excel,
            file_name="drr_summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

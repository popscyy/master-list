from pathlib import Path
from io import BytesIO
from datetime import datetime, timedelta
import re

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_DDR_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\DAILY DIGITAL REPORT\2026\APRIL 2026"
DEFAULT_MERGE_ACCOUNTS_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\MERGED ACCOUNTS\2026\APRIL"

def get_latest_ddr_file(directory: str) -> str:
    """Return the latest .xlsx DDR file in the directory, or empty string if none found."""
    base_dir = Path(directory)
    if not base_dir.exists() or not base_dir.is_dir():
        return ""
    files = list(base_dir.glob("*.xlsx"))
    if not files:
        return ""
    latest = max(files, key=lambda f: f.stat().st_mtime)
    return latest.name

def resolve_server_file(server_dir: str, server_input: str) -> tuple[Path | None, str | None]:
    """Resolve a server file by full path, relative path, or filename search."""
    base_dir = Path(server_dir)
    if not base_dir.exists() or not base_dir.is_dir():
        return None, f"Server folder is not reachable: {base_dir}"

    requested = (server_input or "").strip()
    if not requested:
        return None, "Please provide a server file name or path."

    normalized_requested = requested.replace("/", "\\")
    requested_path = Path(normalized_requested)
    
    if requested_path.exists() and requested_path.is_file():
        return requested_path, None
    
    candidate = base_dir / normalized_requested
    if candidate.exists() and candidate.is_file():
        return candidate, None
    
    requested_name = requested_path.name
    if not requested_name:
        return None, "Please provide a valid file name or relative path."
    
    matches = [p for p in base_dir.rglob("*") if p.is_file() and p.name.lower() == requested_name.lower()]
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        preview = "\n".join(str(p) for p in matches[:5])
        return None, f"Multiple files named '{requested_name}' were found. Please use a relative path.\nMatches:\n{preview}"
    
    return None, f"Server file not found from input: {requested}"

def read_table(file_source, sheet_name="Digital Result"):
    """Read Excel file, specifically looking for 'Digital Result' sheet."""
    if hasattr(file_source, "name"):
        try:
            return pd.read_excel(file_source, sheet_name=sheet_name)
        except:
            return pd.read_excel(file_source)
    else:
        try:
            return pd.read_excel(file_source, sheet_name=sheet_name)
        except:
            return pd.read_excel(file_source)

def load_merge_accounts(file_path) -> tuple[pd.DataFrame, list]:
    debug_lines = []
    try:
        df = pd.read_excel(file_path)
        
        debug_lines.append(("Merge Accounts columns", df.columns.tolist()))
        debug_lines.append(("Merge Accounts shape", df.shape))
        
        account_col = 'ACCOUNT NUMBER'
        placement_col = 'PLACEMENT'
        
        if account_col not in df.columns or placement_col not in df.columns:
            debug_lines.append(("ERROR", f"Expected columns not found. Got: {df.columns.tolist()}"))
            return pd.DataFrame(), debug_lines
        
        result = df[[account_col, placement_col]].drop_duplicates().copy()
        result = result.rename(columns={account_col: 'Account No.'})
        
        debug_lines.append(("Account No. sample", result['Account No.'].head(5).tolist()))
        debug_lines.append(("PLACEMENT sample", result['PLACEMENT'].dropna().unique()[:5].tolist()))
        
        return result, debug_lines
        
    except Exception as e:
        debug_lines.append(("EXCEPTION", str(e)))
        return pd.DataFrame(), debug_lines


def process_outbox_sms_with_placement(outbox_df: pd.DataFrame, merge_accounts: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    debug_lines = []
    
    if 'Account No.' not in outbox_df.columns:
        debug_lines.append(("ERROR", "Account No. column missing from outbox"))
        return outbox_df, debug_lines

    outbox_df = outbox_df.copy()
    merge_accounts = merge_accounts.copy()

    def normalize_account(series):
        return (
            pd.to_numeric(series, errors='coerce')
            .astype('Int64')
            .astype(str)
            .str.strip()
            .str.replace(r'\.0$', '', regex=True)
            .str.replace('<NA>', '', regex=False)
        )

    outbox_df['Account No. Key'] = normalize_account(outbox_df['Account No.'])
    merge_accounts['Account No. Key'] = normalize_account(merge_accounts['Account No.'])

    common = set(outbox_df['Account No. Key']) & set(merge_accounts['Account No. Key'])
    debug_lines.append(("Outbox Account No. Key sample", outbox_df['Account No. Key'].head(5).tolist()))
    debug_lines.append(("Merge Accounts Key sample", merge_accounts['Account No. Key'].head(5).tolist()))
    debug_lines.append(("Matching keys found", len(common)))

    merged = outbox_df.merge(
        merge_accounts[['Account No. Key', 'PLACEMENT']],
        on='Account No. Key',
        how='left'
    )
    merged = merged.drop(columns=['Account No. Key'])
    
    debug_lines.append(("PLACEMENT null count after merge", merged['PLACEMENT'].isna().sum()))
    
    return merged, debug_lines

def clean_and_pivot_sms_placement(df: pd.DataFrame):
    if 'PLACEMENT' not in df.columns or 'Status' not in df.columns:
            st.error(f"Missing columns. Available: {df.columns.tolist()}")
            return df, pd.DataFrame()
        
    # Case-insensitive status filter
    delivered = df[df['Status'].astype(str).str.strip().str.casefold() == 'delivered'].copy()
        
    if delivered.empty:
        st.warning(f"No 'Delivered' records found. Unique Status values: {df['Status'].unique().tolist()}")
        return delivered, pd.DataFrame()
    
    # More robust date column detection
    date_col = None
    for col in delivered.columns:
        col_lower = col.lower().strip()
        if 'submission' in col_lower or ('date' in col_lower and 'time' in col_lower):
            date_col = col
            break
    if not date_col:
        for col in delivered.columns:
            if 'date' in col.lower():
                date_col = col
                break

    """Create pivot table: Delivered SMS count per PLACEMENT with Date breakdown."""
    if 'PLACEMENT' not in df.columns or 'Status' not in df.columns:
        return df, pd.DataFrame()
    
    # Filter for Delivered status
    delivered = df[df['Status'].astype(str).str.strip().str.casefold() == 'delivered'].copy()
    
    if delivered.empty:
        return delivered, pd.DataFrame()
    
    # Identify submission date column (likely 'Submission Date / Time' or similar)
    date_col = None
    for col in delivered.columns:
        if 'submission' in col.lower() or 'date' in col.lower():
            date_col = col
            break
    
    if not date_col:
        date_col = delivered.columns[1] if len(delivered.columns) > 1 else None
    
    # Prepare data for pivot
    pivot_data = delivered[['PLACEMENT', 'Account No.']].copy()
    
    if date_col:
        pivot_data['Submission Date'] = pd.to_datetime(delivered[date_col], errors='coerce').dt.date
        
        # Create pivot with date breakdown
        pivot = pivot_data.groupby(['Submission Date', 'PLACEMENT']).size().reset_index(name='Count of Account No.')
        pivot = pivot.sort_values(['Submission Date', 'PLACEMENT']).reset_index(drop=True)
    else:
        # Create pivot without date breakdown
        pivot = pivot_data.groupby('PLACEMENT').size().reset_index(name='Count of Account No.')
        pivot = pivot.sort_values('PLACEMENT').reset_index(drop=True)
    
    # Calculate grand total
    grand_total = pivot['Count of Account No.'].sum()
    
    if date_col:
        pivot = pd.concat([
            pivot,
            pd.DataFrame({"Submission Date": [None], "PLACEMENT": ["Grand Total"], "Count of Account No.": [grand_total]})
        ], ignore_index=True)
    else:
        pivot = pd.concat([
            pivot,
            pd.DataFrame({"PLACEMENT": ["Grand Total"], "Count of Account No.": [grand_total]})
        ], ignore_index=True)
    
    return delivered, pivot

def to_sms_tracker_excel(cleaned: pd.DataFrame, pivot: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Pivot Table"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    grand_total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    grand_total_font = Font(bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add header row
    header_row = ws1.append(["Row Labels", "Count of SMS Dispo"])
    for cell in ws1[1]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
    
    # Add autofilter
    ws1.auto_filter.ref = "A1:B1"
    
    # Track row numbers for outline levels
    row_num = 2
    outline_level = 1
    
    # Group by Product and create hierarchical output
    grouped = pivot[pivot["Row Labels"] != "Grand Total"].groupby("Row Labels")
    
    for product, group in grouped:
        # Add main category with total
        product_total = group["Count of SMS Dispo"].sum()
        row = ws1.append([product, product_total])
        
        # Style main category row
        for cell in ws1[row_num]:
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row_num += 1
        
        # Add subcategories
        for _, row_data in group.iterrows():
            if row_data["Column"]:  # Only add subcategories with non-empty column names
                row = ws1.append([f"  {row_data['Column']}", row_data["Count of SMS Dispo"]])
                
                # Style subcategory row
                for cell in ws1[row_num]:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left")
                
                row_num += 1
    
    # Add grand total row
    grand_total_row = pivot[pivot["Row Labels"] == "Grand Total"]
    if not grand_total_row.empty:
        row = ws1.append(["Grand Total", grand_total_row["Count of SMS Dispo"].iloc[0]])
        
        # Style grand total row
        for cell in ws1[row_num]:
            cell.fill = grand_total_fill
            cell.font = grand_total_font
            cell.border = thin_border
    
    # Set column widths
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    
    # Add cleaned data sheet
    ws2 = wb.create_sheet("Cleaned Data")
    for r in cleaned.itertuples(index=False):
        ws2.append(r)
    
    # Format cleaned data sheet
    for col_num, col_name in enumerate(cleaned.columns, 1):
        ws2.cell(row=1, column=col_num).font = Font(bold=True)
        ws2.cell(row=1, column=col_num).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def clean_and_pivot_ddr(df: pd.DataFrame):
    """Clean DDR data from 'Digital Result' sheet and create pivot table."""
    
    # Check if DataFrame is empty
    if df.empty:
        st.warning("The uploaded file contains no data.")
        return pd.DataFrame(), pd.DataFrame({"Row Labels": ["Grand Total"], "Column": [""], "Count of SMS Dispo": [0]})
    
    # Check if required columns exist
    required_cols = ['Product', 'Bucket', 'SMS Dispo']
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if missing_cols:
        st.error(f"Missing required columns: {missing_cols}")
        st.write(f"Available columns: {df.columns.tolist()}")
        return pd.DataFrame(), pd.DataFrame({"Row Labels": ["Grand Total"], "Column": [""], "Count of SMS Dispo": [0]})
    
    # Create cleaned dataframe with only needed columns
    cleaned = df[['Product', 'Bucket', 'SMS Dispo']].copy()
    
    # Rename columns for consistency
    cleaned.columns = ['Product', 'Column', 'SMS Dispo']
    
    # Remove rows where SMS Dispo is empty or NaN
    cleaned = cleaned[cleaned['SMS Dispo'].notna() & (cleaned['SMS Dispo'].astype(str).str.strip() != "")]
    
    # Check if cleaned data is empty after filtering
    if cleaned.empty:
        st.warning("No valid SMS Dispo entries found in the file.")
        return cleaned, pd.DataFrame({"Row Labels": ["Grand Total"], "Column": [""], "Count of SMS Dispo": [0]})
    
    try:
        # Create pivot table
        pivot = cleaned.groupby(['Product', 'Column'])['SMS Dispo'].count().reset_index()
        pivot = pivot.rename(columns={'Product': 'Row Labels', 'Column': 'Column', 'SMS Dispo': 'Count of SMS Dispo'})
        
        # Sort properly
        pivot = pivot.sort_values(['Row Labels', 'Column']).reset_index(drop=True)
        
        # Calculate grand total
        grand_total = pivot['Count of SMS Dispo'].sum()
        
        # Add grand total row
        pivot = pd.concat([
            pivot,
            pd.DataFrame({"Row Labels": ["Grand Total"], "Column": [""], "Count of SMS Dispo": [grand_total]})
        ], ignore_index=True)
        
        return cleaned, pivot
        
    except Exception as e:
        st.error(f"Error creating pivot table: {str(e)}")
        return cleaned, pd.DataFrame({"Row Labels": ["Grand Total"], "Column": [""], "Count of SMS Dispo": [0]})
    
# Initialize session state
if "processed_data" not in st.session_state:
    st.session_state.processed_data = None
if "processed_outbox_data" not in st.session_state:
    st.session_state.processed_outbox_data = None

# Streamlit UI
st.title("SMS Blasting Tracker Count")
st.caption("Generate multiple pivot table summaries from SMS tracking data.")

# Initialize tabs for organization
tab1, tab2 = st.tabs(["SMS Blasting Sent Summary", "SMS Blasting Delivered Summary"])

with tab1:
    st.markdown("### Summary 1: SMS Sent Summary Count")
    with st.form("ddr_form"):
        st.markdown("#### Digital Result Sheet File")
        ddr_server_name = st.text_input(
            "Digital Result Sheet file name/path",
            value=get_latest_ddr_file(DEFAULT_DDR_DIR),
            help=f"Server folder: {DEFAULT_DDR_DIR}"
        )
        ddr_upload = st.file_uploader(
            "Or upload Digital Result Sheet",
            type=["xlsx"],
            accept_multiple_files=False,
            key="ddr_upload",
        )
        submitted = st.form_submit_button("Submit", use_container_width=True)

    if submitted:
        ddr_source = None
        ddr_selected_name = None
        errors = []
        
        if ddr_upload is not None:
            ddr_source = ddr_upload
            ddr_selected_name = ddr_upload.name
        else:
            resolved_ddr, ddr_error = resolve_server_file(DEFAULT_DDR_DIR, ddr_server_name)
            if ddr_error:
                errors.append(f"Digital Result Sheet: {ddr_error}")
            else:
                ddr_source = str(resolved_ddr)
                ddr_selected_name = Path(ddr_source).name
        
        if errors:
            st.error("\n".join(errors))
        else:
            try:
                ddr_df = read_table(ddr_source)
                cleaned, pivot = clean_and_pivot_ddr(ddr_df)
                
                st.success("Digital Result Sheet processed and output ready.")

                st.markdown("#### Pivot Table Preview")
                st.dataframe(pivot, use_container_width=True)

                # Filename: sms tracker-<date>.xlsx (date from file name)
                date_match = re.search(r'(\d{8}|\d{6})', ddr_selected_name or "")
                if date_match:
                    token = date_match.group(1)
                    try:
                        if len(token) == 8:
                            base_dt = datetime.strptime(token, "%m%d%Y")
                        else:
                            base_dt = datetime.strptime(token, "%m%d%y")
                    except Exception:
                        base_dt = datetime.now()
                else:
                    base_dt = datetime.now()
                
                yesterday = datetime.now() - timedelta(days=1)
                out_filename = f"sms_history-tracker-{yesterday.strftime('%m%d%y')}.xlsx"
                excel_bytes = to_sms_tracker_excel(cleaned, pivot)
                
                st.download_button(
                    "Download SMS Tracker Output",
                    data=excel_bytes,
                    file_name=out_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                
                st.markdown("#### Cleaned Data Preview")
                st.dataframe(cleaned.head(200), use_container_width=True)
                
            except Exception as exc:
                st.error(f"Processing failed: {exc}")

with tab2:
    st.markdown("### Summary 2: Delivered SMS Blasting Summary Count")
    with st.form("outbox_form"):
        st.markdown("#### Merge Accounts File")
        # Set default merge accounts filename to today's date in the format maya_merged_accounts_MMDDYY.xlsx
        today_str = datetime.now().strftime("%m%d%y")
        default_merge_filename = f"maya_merged_accounts_{today_str}.xlsx"
        merge_server_name = st.text_input(
            "Merge Accounts file name",
            value=default_merge_filename,
            help=f"Server folder: {DEFAULT_MERGE_ACCOUNTS_DIR}",
            key="merge_input"
        )
        merge_upload = st.file_uploader(
            "Or upload Merge Accounts file",
            type=["xlsx"],
            accept_multiple_files=False,
            key="merge_upload",
        )
        
        st.markdown("#### Outbox SMS History File")
        outbox_upload = st.file_uploader(
            "Upload Outbox SMS History file",
            type=["xlsx"],
            accept_multiple_files=False,
            key="outbox_upload",
        )
        
        submitted_outbox = st.form_submit_button("Submit", use_container_width=True)

    if submitted_outbox:
        merge_source = None
        outbox_source = None
        all_debug = []
        errors = []
        
        # Resolve Merge Accounts file
        if merge_upload is not None:
            merge_source = merge_upload
        else:
            resolved_merge, merge_error = resolve_server_file(DEFAULT_MERGE_ACCOUNTS_DIR, merge_server_name)
            if merge_error:
                errors.append(f"Merge Accounts: {merge_error}")
            else:
                merge_source = str(resolved_merge)
        
        # Resolve Outbox file
        if outbox_upload is not None:
            outbox_source = outbox_upload
        else:
            resolved_outbox, outbox_error = resolve_server_file(DEFAULT_MERGE_ACCOUNTS_DIR, outbox_server_name)
            if outbox_error:
                errors.append(f"Outbox SMS History: {outbox_error}")
            else:
                outbox_source = str(resolved_outbox)
        
        if errors:
            st.error("\n".join(errors))
        else:
            try:
                # Load data
                merge_accounts, debug_merge = load_merge_accounts(merge_source)
                all_debug.extend(debug_merge)
                
                outbox_df = read_table(outbox_source)
                
                if merge_accounts.empty:
                    st.error("Failed to load Merge Accounts data.")
                else:
                    # Add PLACEMENT column and create pivot
                    outbox_with_placement, debug_process = process_outbox_sms_with_placement(outbox_df, merge_accounts)
                    all_debug.extend(debug_process)
                    
                    delivered_df, placement_pivot = clean_and_pivot_sms_placement(outbox_with_placement)
                    
                    if placement_pivot.empty:
                        st.warning("No 'Delivered' records found in the data.")
                    else:
                        st.success("Outbox SMS data processed successfully.")
                        
                        st.markdown("#### Delivered SMS Pivot Table")
                        st.dataframe(placement_pivot, use_container_width=True)
                        
                        # Create Excel file for download
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Delivered SMS Blasting Summary"
                        
                        # Add header
                        for col_num, col_name in enumerate(placement_pivot.columns, 1):
                            cell = ws.cell(row=1, column=col_num)
                            cell.value = col_name
                            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Add data
                        for row_num, row_data in enumerate(placement_pivot.values, 2):
                            for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row_num, column=col_num)
                                cell.value = value
                                if 'Grand Total' in str(placement_pivot.iloc[row_num-2, 0]):
                                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                                    cell.font = Font(bold=True)
                        
                        # Set column widths
                        ws.column_dimensions['A'].width = 30
                        ws.column_dimensions['B'].width = 20
                        
                        # Add autofilter
                        ws.auto_filter.ref = f"A1:B{len(placement_pivot) + 1}"
                        
                        excel_output = BytesIO()
                        wb.save(excel_output)
                        excel_output.seek(0)
                        
                        st.download_button(
                            "Download Delivered SMS Report",
                            data=excel_output.getvalue(),
                            file_name="sms_tracker-history-042126.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                        
                        st.session_state.processed_outbox_data = {
                            'delivered_df': delivered_df,
                            'placement_pivot': placement_pivot
                        }
                        
                        st.markdown("#### Preview Records")
                        st.dataframe(delivered_df.head(100), use_container_width=True)
            
            except Exception as exc:
                st.error(f"Processing failed: {exc}")
            
            finally:
                # Always renders at the bottom regardless of success or failure
                if all_debug:
                    with st.expander("🔍 Debug Info", expanded=False):
                        for label, value in all_debug:
                            st.write(f"**{label}:**", value)
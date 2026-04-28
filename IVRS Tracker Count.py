from pathlib import Path
from io import BytesIO
from datetime import datetime, timedelta
import re

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


DEFAULT_IVRS_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\BLASTING\IVRS"
DEFAULT_MERGED_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\MERGED ACCOUNTS\2026\APRIL"

def get_latest_ivrs_file(directory: str) -> str:
	"""Return the latest .xlsx IVRS file in the directory, or empty string if none found."""
	base_dir = Path(directory)
	if not base_dir.exists() or not base_dir.is_dir():
		return ""
	files = list(base_dir.glob("VB-*.xlsx"))
	if not files:
		return ""
	latest = max(files, key=lambda f: f.stat().st_mtime)
	return latest.name

def get_yesterday_merged_file() -> str:
	"""Return the merged file name for yesterday's date in maya_merged_accounts_MMDDYY.xlsx format."""
	yest = (datetime.now() - timedelta(days=1)).strftime("%m%d%y")
	return f"maya_merged_accounts_{yest}.xlsx"

DEFAULT_IVRS_FILE = get_latest_ivrs_file(DEFAULT_IVRS_DIR)
DEFAULT_MERGED_FILE = get_yesterday_merged_file()


def resolve_server_file(server_dir: str, server_input: str) -> tuple[Path | None, str | None]:
	"""Resolve a server file by full path, relative path, or filename search."""
	base_dir = Path(server_dir)
	if not base_dir.exists() or not base_dir.is_dir():
		return None, f"Server folder is not reachable: {base_dir}"

	requested = (server_input or "").strip()
	if not requested:
		return None, "Please provide a server file name or path."

	normalized_requested = requested.replace("/", "\\")
	requested_path = Path(normalized_requested)

	if requested_path.exists() and requested_path.is_file():
		return requested_path, None

	candidate = base_dir / normalized_requested
	if candidate.exists() and candidate.is_file():
		return candidate, None

	requested_name = requested_path.name
	if not requested_name:
		return None, "Please provide a valid file name or relative path."

	matches = [
		p for p in base_dir.rglob("*") if p.is_file() and p.name.lower() == requested_name.lower()
	]
	if len(matches) == 1:
		return matches[0], None
	if len(matches) > 1:
		preview = "\n".join(str(p) for p in matches[:5])
		return None, (
			f"Multiple files named '{requested_name}' were found. "
			"Please use a relative path.\n"
			f"Matches:\n{preview}"
		)

	return None, f"Server file not found from input: {requested}"


def read_table(file_source) -> pd.DataFrame:
	"""Read CSV or Excel from either uploaded file or filesystem path."""
	if hasattr(file_source, "name"):
		name = str(file_source.name)
		ext = Path(name).suffix.lower()
		if ext == ".csv":
			return pd.read_csv(file_source)
		return pd.read_excel(file_source)

	path = Path(file_source)
	ext = path.suffix.lower()
	if ext == ".csv":
		return pd.read_csv(path)
	return pd.read_excel(path)


def to_account_key(value) -> str:
	"""Normalize account numbers for robust matching between files."""
	if value is None or pd.isna(value):
		return ""

	if isinstance(value, (int, float)) and not pd.isna(value):
		try:
			return str(int(float(value)))
		except Exception:
			pass

	txt = str(value).strip()
	if txt.lower() in {"", "none", "nan", "null", "na", "n/a"}:
		return ""

	txt = txt.replace(",", "").replace(" ", "")
	parsed = pd.to_numeric(pd.Series([txt]), errors="coerce").iloc[0]
	if not pd.isna(parsed):
		try:
			return str(int(float(parsed)))
		except Exception:
			pass

	digits = "".join(ch for ch in txt if ch.isdigit())
	if digits:
		return digits

	return txt


def pick_column(columns: list[str], candidates: list[str]) -> str | None:
	"""Pick first matching column from candidates using uppercase normalized names."""
	upper_map = {str(col).strip().upper(): col for col in columns}
	for candidate in candidates:
		found = upper_map.get(candidate.upper())
		if found is not None:
			return found
	return None


def build_ivrs_tracker(ivrs_df: pd.DataFrame, merged_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
	"""Build transformed IVRS table and pivot count by placement."""
	ivrs_account_col = pick_column(
		ivrs_df.columns.tolist(),
		["ACCOUNT NO.", "ACCOUNT NO", "ACCOUNT NUMBER", "ACCOUNT_NO", "ACCOUNT_NUMBER", "ACCOUNT"],
	)
	if ivrs_account_col is None:
		raise ValueError("IVRS file is missing Account No./Account Number column.")

	merged_account_col = pick_column(
		merged_df.columns.tolist(),
		["ACCOUNT NUMBER", "ACCOUNT NO.", "ACCOUNT NO", "ACCOUNT_NO", "ACCOUNT_NUMBER", "ACCOUNT"],
	)
	merged_placement_col = pick_column(
		merged_df.columns.tolist(),
		["PLACEMENT", "AGENCY", "AGENCY_NAME"],
	)
	if merged_account_col is None or merged_placement_col is None:
		raise ValueError("Merged file must contain both account and PLACEMENT columns.")

	ivrs_base = ivrs_df[[ivrs_account_col]].copy()
	ivrs_base.columns = ["Account No."]
	ivrs_base["_ACCOUNT_KEY"] = ivrs_base["Account No."].map(to_account_key)

	merged_base = merged_df[[merged_account_col, merged_placement_col]].copy()
	merged_base.columns = ["_MERGED_ACCOUNT", "PLACEMENT"]
	merged_base["_ACCOUNT_KEY"] = merged_base["_MERGED_ACCOUNT"].map(to_account_key)
	merged_base = merged_base[merged_base["_ACCOUNT_KEY"] != ""].copy()
	merged_base = merged_base.drop_duplicates(subset=["_ACCOUNT_KEY"], keep="first")

	final_ivrs = ivrs_base.merge(
		merged_base[["_ACCOUNT_KEY", "PLACEMENT"]],
		on="_ACCOUNT_KEY",
		how="left",
	)
	final_ivrs = final_ivrs[["Account No.", "PLACEMENT"]]

	pivot_df = (
		final_ivrs.groupby("PLACEMENT", dropna=False)["Account No."]
		.count()
		.reset_index(name="Count of Account No.")
	)
	pivot_df["PLACEMENT"] = pivot_df["PLACEMENT"].fillna("").astype(str).str.strip()
	pivot_df = pivot_df[(pivot_df["PLACEMENT"] != "") & (pivot_df["Count of Account No."] > 0)].copy()
	pivot_df = pivot_df.sort_values("PLACEMENT").reset_index(drop=True)
	pivot_df = pivot_df.rename(columns={"PLACEMENT": "Row Labels"})

	grand_total = int(pivot_df["Count of Account No."].sum())
	pivot_df = pd.concat(
		[
			pivot_df,
			pd.DataFrame([{"Row Labels": "Grand Total", "Count of Account No.": grand_total}]),
		],
		ignore_index=True,
	)

	return final_ivrs, pivot_df


def to_excel_bytes(ivrs_final: pd.DataFrame, pivot_df: pd.DataFrame) -> bytes:
	"""Write pivot-style output workbook similar to the sample screenshot."""
	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = "Pivot Count"

	# Leave the top two rows blank and start headers at row 3.
	header_row = 3
	worksheet.cell(row=header_row, column=1, value="Row Labels")
	worksheet.cell(row=header_row, column=2, value="Count of Account No.")

	header_fill = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
	header_font = Font(bold=True)
	for col_idx in (1, 2):
		cell = worksheet.cell(row=header_row, column=col_idx)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right")

	start_row = header_row + 1
	for offset, (_, row) in enumerate(pivot_df.iterrows(), start=0):
		excel_row = start_row + offset
		label = row["Row Labels"]
		count_value = int(row["Count of Account No."])
		worksheet.cell(row=excel_row, column=1, value=label)
		worksheet.cell(row=excel_row, column=2, value=count_value)
		worksheet.cell(row=excel_row, column=2).number_format = "0"

		if str(label) == "Grand Total":
			for col_idx in (1, 2):
				cell = worksheet.cell(row=excel_row, column=col_idx)
				cell.fill = header_fill
				cell.font = Font(bold=True)

	# Set widths similar to the sample.
	worksheet.column_dimensions["A"].width = 34
	worksheet.column_dimensions["B"].width = 20

	output = BytesIO()
	workbook.save(output)
	output.seek(0)
	return output.getvalue()


def derive_output_filename(merged_file_name: str | None) -> str:
	"""Build output name as ivrs tracker-MMDDYY.xlsx using merged file date + 1 day."""
	base_dt = datetime.now()
	name = str(merged_file_name or "")
	match = re.search(r"(\d{8}|\d{6})", name)
	if match:
		token = match.group(1)
		try:
			if len(token) == 8:
				base_dt = datetime.strptime(token, "%m%d%Y")
			else:
				base_dt = datetime.strptime(token, "%m%d%y")
		except ValueError:
			pass

	ivrs_dt = base_dt + timedelta(days=1)
	return f"ivrs tracker-{ivrs_dt.strftime('%m%d%y')}.xlsx"


st.title("IVRS Tracker Count")
st.caption("Input loader for IVRS and Merged Accounts files.")

if "ivrs_input_path" not in st.session_state:
	st.session_state.ivrs_input_path = None
if "merged_input_path" not in st.session_state:
	st.session_state.merged_input_path = None
if "ivrs_output_df" not in st.session_state:
	st.session_state.ivrs_output_df = None
if "pivot_output_df" not in st.session_state:
	st.session_state.pivot_output_df = None
if "output_excel_bytes" not in st.session_state:
	st.session_state.output_excel_bytes = None
if "output_file_name" not in st.session_state:
	st.session_state.output_file_name = "ivrs tracker.xlsx"

st.subheader("Inputer")

with st.form("ivrs_input_form"):
	st.markdown("### IVRS File")
	ivrs_server_name = st.text_input(
		"IVRS server file name/path",
		value=DEFAULT_IVRS_FILE,
		help=f"Server folder: {DEFAULT_IVRS_DIR}"
	)
	ivrs_upload = st.file_uploader(
		"Or upload IVRS file",
		type=["xlsx", "xls", "xlsb", "csv"],
		accept_multiple_files=False,
		key="ivrs_upload",
	)

	st.markdown("### Merged Accounts File")
	merged_server_name = st.text_input(
		"Merged server file name/path",
		value=DEFAULT_MERGED_FILE,
		help=f"Server folder: {DEFAULT_MERGED_DIR}"
	)
	merged_upload = st.file_uploader(
		"Or upload merged accounts file",
		type=["xlsx", "xls", "xlsb", "csv"],
		accept_multiple_files=False,
		key="merged_upload",
	)

	submitted = st.form_submit_button("Load Inputs", use_container_width=True)

if submitted:
	ivrs_source = None
	merged_source = None
	merged_selected_name = None
	errors: list[str] = []

	if ivrs_upload is not None:
		ivrs_source = f"Uploaded: {ivrs_upload.name}"
	else:
		resolved_ivrs, ivrs_error = resolve_server_file(DEFAULT_IVRS_DIR, ivrs_server_name)
		if ivrs_error:
			errors.append(f"IVRS: {ivrs_error}")
		else:
			ivrs_source = str(resolved_ivrs)

	if merged_upload is not None:
		merged_source = f"Uploaded: {merged_upload.name}"
		merged_selected_name = merged_upload.name
	else:
		resolved_merged, merged_error = resolve_server_file(DEFAULT_MERGED_DIR, merged_server_name)
		if merged_error:
			errors.append(f"Merged Accounts: {merged_error}")
		else:
			merged_source = str(resolved_merged)
			merged_selected_name = resolved_merged.name

	if errors:
		st.error("\n".join(errors))
	else:
		st.session_state.ivrs_input_path = ivrs_source
		st.session_state.merged_input_path = merged_source

		try:
			ivrs_df = read_table(ivrs_upload if ivrs_upload is not None else Path(ivrs_source))
			merged_df = read_table(merged_upload if merged_upload is not None else Path(merged_source))

			ivrs_final_df, pivot_df = build_ivrs_tracker(ivrs_df, merged_df)
			st.session_state.ivrs_output_df = ivrs_final_df
			st.session_state.pivot_output_df = pivot_df
			st.session_state.output_excel_bytes = to_excel_bytes(ivrs_final_df, pivot_df)
			st.session_state.output_file_name = derive_output_filename(merged_selected_name)
			st.success("Inputs loaded and IVRS tracker output generated.")
		except Exception as exc:
			st.error(f"Processing failed: {exc}")

if st.session_state.pivot_output_df is not None:
	st.markdown("### Pivot Count Preview")
	st.dataframe(st.session_state.pivot_output_df, use_container_width=True)

if st.session_state.output_excel_bytes is not None:
	st.download_button(
		"Download IVRS Tracker Output",
		data=st.session_state.output_excel_bytes,
		file_name=st.session_state.output_file_name,
		mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		use_container_width=True,
	)

if st.session_state.ivrs_input_path or st.session_state.merged_input_path:
	st.markdown("### Loaded Inputs")
	st.write(f"IVRS File: {st.session_state.ivrs_input_path or '-'}")
	st.write(f"Merged Accounts File: {st.session_state.merged_input_path or '-'}")

if st.session_state.ivrs_output_df is not None:
	st.markdown("### Final IVRS Structure")
	st.dataframe(st.session_state.ivrs_output_df.head(200), use_container_width=True)



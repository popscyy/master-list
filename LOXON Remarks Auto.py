from io import BytesIO
from pathlib import Path
from datetime import date, datetime, timedelta, time
import json
import re

import msoffcrypto
import pandas as pd
import polars as pl
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from resources.excel_tools import cast_columns, save_xlsx


DEFAULT_WORKBOOK_PASSWORD = "Maya@2026"
DEFAULT_SERVER_MASTERFILE_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\ENDO\MASTERFILE"
DEFAULT_SERVER_ENDORSEMENT_FILE = f"MAYA ENDORSEMENT {date.today():%m%d%Y}.xlsx"



def resolve_server_endorsement_file(server_input: str) -> tuple[Path | None, str | None]:
	"""[STEP 4] Resolve endorsement file from server network path with fallback search."""
	master_dir = Path(DEFAULT_SERVER_MASTERFILE_DIR)
	if not master_dir.exists() or not master_dir.is_dir():
		return None, f"Server folder is not reachable: {master_dir}"

	requested = (server_input or "").strip()
	if not requested:
		return None, "Server file mode is enabled but no file name/path was provided."

	normalized_requested = requested.replace("/", "\\")
	requested_path = Path(normalized_requested)

	# Allow a full file path directly.
	if requested_path.exists() and requested_path.is_file():
		return requested_path, None

	# Allow a path relative to MASTERFILE (e.g. MARCH 2026\file.xlsx).
	candidate = master_dir / normalized_requested
	if candidate.exists() and candidate.is_file():
		return candidate, None

	# Fallback: search by filename anywhere under MASTERFILE.
	requested_name = requested_path.name
	if not requested_name:
		return None, "Please provide a valid server file name or relative path."

	matches = [
		p for p in master_dir.rglob("*") if p.is_file() and p.name.lower() == requested_name.lower()
	]
	if len(matches) == 1:
		return matches[0], None
	if len(matches) > 1:
		preview = "\n".join(str(p) for p in matches[:5])
		return None, (
			f"Multiple files named '{requested_name}' were found under MASTERFILE. "
			"Please paste a relative path, for example 'MARCH 2026\\file.xlsx'.\n"
			f"Matches:\n{preview}"
		)

	return None, f"Server file not found from input: {requested}"


def standardize_column_name(name: str) -> str:
	"""[UTILITY] Normalize column names to lowercase with underscores for consistent matching across all steps."""
	cleaned = "".join(ch if ch.isalnum() else "_" for ch in str(name).strip().lower())
	while "__" in cleaned:
		cleaned = cleaned.replace("__", "_")
	return cleaned.strip("_")


def pick_column(columns: list[str], candidates: list[str]) -> str | None:
	"""[UTILITY] Find the first matching column name from a list of candidates (case-insensitive matching)."""
	for candidate in candidates:
		if candidate in columns:
			return candidate
	return None


def to_account_key(value) -> str:
	"""[UTILITY] Normalize account numbers to standardized format for key matching (used in fill missing values)."""
	if value is None or pd.isna(value):
		return ""

	# Fast-path numeric values, including floats read from Excel.
	if isinstance(value, (int, float)) and not pd.isna(value):
		try:
			return str(int(float(value)))
		except Exception:
			pass

	txt = str(value).strip()
	if txt.lower() in {"", "none", "nan", "null", "na", "n/a"}:
		return ""

	# Remove common visual separators before parsing.
	txt = txt.replace(",", "").replace(" ", "")

	# Handle scientific notation and decimal-text numbers.
	parsed = pd.to_numeric(pd.Series([txt]), errors="coerce").iloc[0]
	if not pd.isna(parsed):
		try:
			return str(int(float(parsed)))
		except Exception:
			pass

	# Fallback: keep only digits when the value is mostly numeric text.
	digits = "".join(ch for ch in txt if ch.isdigit())
	if digits:
		return digits

	return txt


def is_missing_like(value) -> bool:
	"""[UTILITY] Check if a value is considered missing/null across multiple representation formats."""
	if value is None or pd.isna(value):
		return True
	txt = str(value).strip().lower()
	return txt in {"", "none", "nan", "null", "na", "n/a"}


def normalize_number_contacted(value):
	"""[UTILITY] Normalize phone numbers to 63XxxxxxxxxX format (used in fill missing values and Step 5)."""
	if value is None or pd.isna(value):
		return value

	txt = str(value).strip()
	if txt == "":
		return value

	digits = "".join(ch for ch in txt if ch.isdigit())
	if digits == "":
		return txt

	if digits.startswith("63"):
		return digits
	return "63" + digits[-10:]


def _read_excel_bytes(file_bytes: bytes, ext: str, sheet_name=0) -> pd.DataFrame:
	"""[UTILITY] Read Excel/CSV bytes into pandas DataFrame with fallback engine support."""
	if ext == ".csv":
		return pd.read_csv(BytesIO(file_bytes))

	engines: list[str | None] = [None]
	if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
		engines = ["openpyxl", None]
	elif ext == ".xlsb":
		engines = ["pyxlsb", "openpyxl", None]
	elif ext == ".xls":
		engines = ["xlrd", "openpyxl", None]

	last_error = None
	for engine in engines:
		try:
			# Read with all columns as string to preserve '+' signs and leading zeros in phone numbers
			df = pd.read_excel(BytesIO(file_bytes), engine=engine, sheet_name=sheet_name, dtype=str)
			return df
		except Exception as exc:
			last_error = exc

	raise ValueError(f"Unable to read workbook using available engines: {last_error}")


def _decrypt_excel_bytes(file_bytes: bytes, workbook_password: str) -> bytes:
	"""[UTILITY] Decrypt password-protected Excel workbooks using msoffcrypto (used for Step 4 endorsement file)."""
	decrypted = BytesIO()
	office_file = msoffcrypto.OfficeFile(BytesIO(file_bytes))
	office_file.load_key(password=workbook_password)
	office_file.decrypt(decrypted)
	decrypted.seek(0)
	return decrypted.getvalue()


def _file_obj_to_bytes(file_obj) -> bytes:
	"""[UTILITY] Convert file-like object or Path to bytes for consistent file handling."""
	if isinstance(file_obj, Path):
		return file_obj.read_bytes()
	if hasattr(file_obj, "getvalue"):
		return file_obj.getvalue()
	if hasattr(file_obj, "read"):
		return file_obj.read()
	raise TypeError(f"Unsupported file object: {type(file_obj)!r}")


def _load_workbook_bytes(file_obj, workbook_password: str | None = None):
	"""[UTILITY] Load Excel workbook from file object with automatic decryption if needed."""
	ext = Path(file_obj.name).suffix.lower()
	file_bytes = _file_obj_to_bytes(file_obj)
	effective_password = workbook_password or DEFAULT_WORKBOOK_PASSWORD

	if ext == ".csv":
		raise ValueError("CSV files cannot preserve workbook formatting.")

	try:
		return load_workbook(BytesIO(file_bytes))
	except Exception as first_error:
		if not effective_password:
			raise ValueError(str(first_error))
		try:
			decrypted_bytes = _decrypt_excel_bytes(file_bytes, effective_password)
			return load_workbook(BytesIO(decrypted_bytes))
		except Exception as second_error:
			raise ValueError(f"{first_error}. Decryption retry failed: {second_error}")


def read_endorsement_pout(file_obj, workbook_password: str | None = None) -> pd.DataFrame:
	"""[STEP 4] Read and extract POUT sheet from encrypted endorsement workbook."""
	file_bytes = _file_obj_to_bytes(file_obj)
	ext = Path(file_obj.name).suffix.lower() if hasattr(file_obj, "name") else Path(file_obj).suffix.lower()
	effective_password = workbook_password or DEFAULT_WORKBOOK_PASSWORD

	if ext == ".csv":
		return _read_excel_bytes(file_bytes, ext)

	# Load all sheets first so we can match POUT case-insensitively.
	try:
		sheets = _read_excel_bytes(file_bytes, ext, sheet_name=None)
	except Exception as first_error:
		if not effective_password:
			raise ValueError(str(first_error))
		try:
			decrypted_bytes = _decrypt_excel_bytes(file_bytes, effective_password)
			sheets = _read_excel_bytes(decrypted_bytes, ext, sheet_name=None)
		except Exception as second_error:
			raise ValueError(f"{first_error}. Decryption retry failed: {second_error}")

	if isinstance(sheets, dict):
		for name, df in sheets.items():
			if str(name).strip().upper() == "POUT":
				return df
		raise ValueError("Endorsement file does not contain a POUT sheet.")

	# Single-sheet workbook fallback
	return sheets


def read_endorsement_active(file_obj, workbook_password: str | None = None) -> pd.DataFrame:
	"""[FILL MISSING VALUES] Read and extract ACTIVE sheet from encrypted endorsement workbook (for number_contacted lookup)."""
	file_bytes = _file_obj_to_bytes(file_obj)
	ext = Path(file_obj.name).suffix.lower() if hasattr(file_obj, "name") else Path(file_obj).suffix.lower()
	effective_password = workbook_password or DEFAULT_WORKBOOK_PASSWORD

	if ext == ".csv":
		return _read_excel_bytes(file_bytes, ext)

	# Load all sheets first so we can match ACTIVE case-insensitively.
	try:
		sheets = _read_excel_bytes(file_bytes, ext, sheet_name=None)
	except Exception as first_error:
		if not effective_password:
			raise ValueError(str(first_error))
		try:
			decrypted_bytes = _decrypt_excel_bytes(file_bytes, effective_password)
			sheets = _read_excel_bytes(decrypted_bytes, ext, sheet_name=None)
		except Exception as second_error:
			raise ValueError(f"{first_error}. Decryption retry failed: {second_error}")

	if isinstance(sheets, dict):
		for name, df in sheets.items():
			if str(name).strip().upper() == "ACTIVE":
				return df
		raise ValueError("Endorsement file does not contain an ACTIVE sheet.")

	# Single-sheet workbook fallback
	return sheets


def read_excel(file_obj, workbook_password: str | None = None) -> pd.DataFrame:
	"""[UTILITY] Read Excel/CSV file with optional password support (used across all steps for file upload)."""
	file_bytes = _file_obj_to_bytes(file_obj)
	ext = Path(file_obj.name).suffix.lower() if hasattr(file_obj, "name") else Path(file_obj).suffix.lower()
	effective_password = workbook_password or DEFAULT_WORKBOOK_PASSWORD

	try:
		df = _read_excel_bytes(file_bytes, ext)
	except Exception as first_error:
		if not effective_password or ext == ".csv":
			raise ValueError(str(first_error))

		try:
			decrypted_bytes = _decrypt_excel_bytes(file_bytes, effective_password)
			df = _read_excel_bytes(decrypted_bytes, ext)
		except Exception as second_error:
			raise ValueError(f"{first_error}. Decryption retry failed: {second_error}")

	return df


def to_output_excel(df: pd.DataFrame) -> bytes:
	"""[UTILITY] Convert DataFrame to Excel bytes with account formatting (legacy, mostly replaced by preserve version)."""
	output = BytesIO()
	output_df = df.copy()
	account_col = pick_column(list(output_df.columns), ["account_number", "account_no", "account", "account_num"])
	if account_col is not None:
		output_df[account_col] = output_df[account_col].map(lambda value: "" if is_missing_like(value) else to_account_key(value))
	with pd.ExcelWriter(output, engine="openpyxl") as writer:
		output_df.to_excel(writer, index=False, sheet_name="MADRID Feedback")
		ws = writer.sheets["MADRID Feedback"]
		for header_cell in ws[1]:
			header_cell.alignment = Alignment(horizontal="center", vertical="center")
			header_cell.font = Font(bold=True)
		for col_idx, col_name in enumerate(output_df.columns, start=1):
			max_len = len(str(col_name))
			for value in output_df[col_name].head(5000):
				val_len = len(str(value)) if value is not None else 0
				if val_len > max_len:
					max_len = val_len
			ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
	output.seek(0)
	return output.getvalue()


def to_output_excel_raw(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
	"""[UTILITY] Convert DataFrame to Excel bytes without any data or format transformation."""
	output = BytesIO()
	with pd.ExcelWriter(output, engine="openpyxl") as writer:
		df.to_excel(writer, index=False, sheet_name=sheet_name)
	output.seek(0)
	return output.getvalue()


def to_output_excel_preserve(df: pd.DataFrame, sheet_name: str = "Combined") -> bytes:
	"""[STEP 1] Convert Step 1 combined DataFrame to Excel bytes with date formatting preserved (mm/dd/yyyy)."""
	output = BytesIO()
	output_df = df.copy()

	# Convert date-like columns to datetime so Excel can apply date formatting reliably.
	for col_name in output_df.columns:
		std_col_name = standardize_column_name(col_name)
		if std_col_name in {"date", "ptp_date", "claim_paid_date", "field_visit_date", "next_call"}:
			parsed = pd.to_datetime(output_df[col_name], errors="coerce")
			if parsed.notna().any():
				output_df[col_name] = parsed
		if std_col_name in {"dialed_number"}:
			# Preserve literal text representation of phone numbers (+63 / leading 0).
			output_df[col_name] = output_df[col_name].map(
				lambda value: None
				if is_missing_like(value)
				# Keep raw string values exactly as-is so '+' remains when present in source data.
				else (
					value
					if isinstance(value, str)
					else (str(int(float(value))) if isinstance(value, float) and value.is_integer() else str(value))
				)
			)

	with pd.ExcelWriter(output, engine="openpyxl") as writer:
		output_df.to_excel(writer, index=False, sheet_name=sheet_name)
		ws = writer.sheets[sheet_name]
		for header_cell in ws[1]:
			header_cell.alignment = Alignment(horizontal="center", vertical="center")
			header_cell.font = Font(bold=True)
		integer_id_col_indices = []
		decimal_col_indices = []
		date_col_indices = []
		text_col_indices = []
		for col_idx, col_name in enumerate(output_df.columns, start=1):
			std_col_name = standardize_column_name(col_name)
			if std_col_name in {"account_no", "account_number", "card_no", "service_no", "s_no"}:
				integer_id_col_indices.append(col_idx)
			if std_col_name in {"ptp_amount", "claim_paid_amount", "balance", "ob"}:
				decimal_col_indices.append(col_idx)
			if std_col_name in {"date", "ptp_date", "claim_paid_date", "field_visit_date", "next_call"}:
				date_col_indices.append(col_idx)
			if std_col_name in {"dialed_number"}:
				# Keep Dialed Number as text to preserve +63/leading 0 and avoid E+11 display.
				text_col_indices.append(col_idx)

			max_len = len(str(col_name))
			for value in output_df[col_name].head(5000):
				val_len = len(str(value)) if value is not None else 0
				if val_len > max_len:
					max_len = val_len
			ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

		for col_idx in integer_id_col_indices:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=col_idx)
				if cell.value is not None and str(cell.value).strip() != "":
					cell.number_format = "0"

		for col_idx in decimal_col_indices:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=col_idx)
				if cell.value is not None and str(cell.value).strip() != "":
					cell.number_format = "0.00"

		for col_idx in date_col_indices:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=col_idx)
				if cell.value is not None and str(cell.value).strip() != "":
					cell.number_format = "dd-mm-yyyy"
					cell.alignment = Alignment(horizontal="left")

		for col_idx in text_col_indices:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=col_idx)
				if cell.value is not None and str(cell.value).strip() != "":
					cell.number_format = "@"
					cell.alignment = Alignment(horizontal="left")

		# Keep Column E left-aligned to match manual Step 1 format.
		if ws.max_column >= 5:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=5)
				if cell.value is not None and str(cell.value).strip() != "":
					cell.alignment = Alignment(horizontal="left")

		# Keep Column H left-aligned to match manual Step 1 format.
		if ws.max_column >= 8:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=8)
				if cell.value is not None and str(cell.value).strip() != "":
					cell.alignment = Alignment(horizontal="left")

		# Keep Column V right-aligned to match manual Step 1 format.
		if ws.max_column >= 22:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=22)
				if cell.value is not None and str(cell.value).strip() != "":
					cell.alignment = Alignment(horizontal="right")

		# Keep Column Y right-aligned to match manual Step 1 format.
		if ws.max_column >= 25:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=25)
				if cell.value is not None and str(cell.value).strip() != "":
					cell.alignment = Alignment(horizontal="right")

		# Column AB should be two-decimal numeric format (0.00).
		if ws.max_column >= 28:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=28)
				if cell.value is not None and str(cell.value).strip() != "":
					cell.number_format = "0.00"

		# Keep specified columns right-aligned to match manual Step 1 format.
		for fixed_right_col in [29, 35, 36, 38, 39]:  # AC, AI, AJ, AL, AM
			if ws.max_column >= fixed_right_col:
				for row_idx in range(2, ws.max_row + 1):
					cell = ws.cell(row=row_idx, column=fixed_right_col)
					if cell.value is not None and str(cell.value).strip() != "":
						cell.alignment = Alignment(horizontal="right")
	output.seek(0)
	return output.getvalue()


def align_to_base_columns(df: pd.DataFrame, base_columns: list[str]) -> pd.DataFrame:
	"""[STEP 1] Align file columns to base file columns (used in combine_three_files_by_header for file 2 and 3)."""
	std_to_actual = {standardize_column_name(col): col for col in df.columns}
	aligned_data = {}
	for base_col in base_columns:
		matched_col = std_to_actual.get(standardize_column_name(base_col))
		if matched_col is not None:
			aligned_data[base_col] = df[matched_col]
		else:
			aligned_data[base_col] = [None] * len(df)
	return pd.DataFrame(aligned_data)


def drop_trailing_empty_sno_rows(df: pd.DataFrame) -> pd.DataFrame:
	"""[STEP 1] Remove trailing rows that only contain S.No values (spreadsheet artifacts)."""
	if df.empty:
		return df

	std_to_actual = {standardize_column_name(col): col for col in df.columns}
	s_no_col = std_to_actual.get("s_no")
	if s_no_col is None:
		return df

	cleaned = df.copy()
	while not cleaned.empty:
		last_row = cleaned.iloc[-1]
		s_no_has_value = not is_missing_like(last_row.get(s_no_col))
		has_other_values = any(
			not is_missing_like(last_row.get(col))
			for col in cleaned.columns
			if col != s_no_col
		)
		if s_no_has_value and not has_other_values:
			cleaned = cleaned.iloc[:-1].copy()
		else:
			break

	return cleaned


def normalize_account_no_integer(df: pd.DataFrame) -> pd.DataFrame:
	"""[STEP 1] Cast Account No column to integer type for consistent numeric formatting."""
	std_to_actual = {standardize_column_name(col): col for col in df.columns}
	account_col = std_to_actual.get("account_no")
	if account_col is None:
		account_col = std_to_actual.get("account_number")
	if account_col is None:
		return df

	cleaned = df.copy()
	numeric = pd.to_numeric(cleaned[account_col], errors="coerce")
	cleaned[account_col] = numeric.round(0).astype("Int64")
	return cleaned


def normalize_step1_amount_text(df: pd.DataFrame) -> pd.DataFrame:
	"""[STEP 1] Normalize amount-like fields to two-decimal text, matching manual combined output."""
	cleaned = df.copy()
	std_to_actual = {standardize_column_name(col): col for col in cleaned.columns}
	amount_cols = {
		"ptp_amount",
		"claim_paid_amount",
		"over_limit_amount",
		"min_payment",
		"monthly_installment",
		"30_days",
	}

	for std_col in amount_cols:
		actual_col = std_to_actual.get(std_col)
		if actual_col is None:
			continue

		def _fmt(value):
			if is_missing_like(value):
				return None
			txt = str(value).strip()
			num = pd.to_numeric([txt.replace(",", "")], errors="coerce")[0]
			if not pd.isna(num):
				return f"{float(num):.2f}"
			return txt

		cleaned[actual_col] = cleaned[actual_col].map(_fmt)

	return cleaned


def combine_three_files_by_header(
	file_1,
	file_2,
	file_3,
	workbook_password: str | None = None,
) -> pd.DataFrame:
	"""[STEP 1] Combine three daily remark files by aligning to base file headers and concatenating rows."""
	df1 = normalize_account_no_integer(drop_trailing_empty_sno_rows(read_excel(file_1, workbook_password=workbook_password)))
	df2 = normalize_account_no_integer(drop_trailing_empty_sno_rows(read_excel(file_2, workbook_password=workbook_password)))
	df3 = normalize_account_no_integer(drop_trailing_empty_sno_rows(read_excel(file_3, workbook_password=workbook_password)))

	base_columns = list(df1.columns)
	aligned_df1 = df1[base_columns].copy()
	aligned_df2 = align_to_base_columns(df2, base_columns)
	aligned_df3 = align_to_base_columns(df3, base_columns)

	combined = pd.concat([aligned_df1, aligned_df2, aligned_df3], ignore_index=True)
	combined = normalize_account_no_integer(combined)
	return normalize_step1_amount_text(combined)


def resolve_report_date_token(df: pd.DataFrame) -> str:
	"""[UTILITY] Extract maximum date from DataFrame for report date token (MMDDYYYY format)."""
	date_candidates = [
		"Date",
		"date",
		"event_datetime_pht",
		"datalate_processed_ts_pht",
		"call_start",
	]
	for col in date_candidates:
		if col not in df.columns:
			continue
		parsed = pd.to_datetime(df[col], errors="coerce")
		parsed = parsed.dropna()
		if not parsed.empty:
			return parsed.max().strftime("%m%d%Y")
	return datetime.now().strftime("%m%d%Y")


def polars_to_excel_bytes(
	df: pl.DataFrame,
	dtype_formats: dict | None = None,
	left_align_col_indices: list[int] | None = None,
) -> bytes:
	"""[UTILITY] Export Polars DataFrame to Excel bytes with dtype formatting (used in Step 2, Step 5 output)."""
	buffer = BytesIO()
	kwargs = {"autofit": True}
	if dtype_formats is not None:
		kwargs["dtype_formats"] = dtype_formats
	df.write_excel(buffer, **kwargs)

	if left_align_col_indices:
		buffer.seek(0)
		workbook = load_workbook(buffer)
		worksheet = workbook.active
		for col_idx in left_align_col_indices:
			if col_idx > worksheet.max_column:
				continue
			for row_idx in range(2, worksheet.max_row + 1):
				cell = worksheet.cell(row=row_idx, column=col_idx)
				if cell.value is not None and str(cell.value).strip() != "":
					cell.alignment = Alignment(horizontal="left")

		formatted_output = BytesIO()
		workbook.save(formatted_output)
		formatted_output.seek(0)
		return formatted_output.getvalue()

	buffer.seek(0)
	return buffer.getvalue()


def format_date_token(value) -> str:
	"""[UTILITY] Format date value to YYYY-MM-DD ISO format for filenames (used in Steps 2-5 output naming)."""
	if value is None:
		return datetime.now().strftime("%Y-%m-%d")
	if isinstance(value, datetime):
		return value.strftime("%Y-%m-%d")
	if isinstance(value, date):
		return value.strftime("%Y-%m-%d")
	parsed = pd.to_datetime([value], errors="coerce")
	if not parsed.isna().all():
		return parsed[0].strftime("%Y-%m-%d")
	return datetime.now().strftime("%Y-%m-%d")


def map_step3_call_type(value):
	"""[STEP 3] Map Remark Type values to Call Type values (Outgoing->Manual, Follow Up->Predictive)."""
	normalized = re.sub(r"\s+", " ", str(value).strip().lower())
	normalized = normalized.replace("_", " ").replace("-", " ")
	normalized = re.sub(r"\s+", " ", normalized).strip()
	if normalized == "outgoing":
		return "Manual"
	if normalized == "follow up":
		return "Predictive"
	return value


def add_step3_call_type_preserve_workbook(file_obj, workbook_password: str | None = None) -> bytes:
	"""[STEP 3] Add Call Type column at position 0, populate from Remark Type, save cleanly."""
	# Read file as DataFrame
	df = read_excel(file_obj, workbook_password=workbook_password)

	# Add Call Type column at position 0 using existing function
	output_df = build_step3_call_type_file(df)

	# Write back to Excel
	output = BytesIO()
	with pd.ExcelWriter(output, engine="openpyxl") as writer:
		output_df.to_excel(writer, index=False, sheet_name="Call Logs")
		ws = writer.sheets["Call Logs"]

		# Force column C to true Excel date values with dd/mm/yyyy display.
		date_col_idx = 3  # C
		if date_col_idx <= ws.max_column:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=date_col_idx)
				if cell.value is None or str(cell.value).strip() == "":
					continue
				parsed = pd.to_datetime([cell.value], errors="coerce", dayfirst=True)[0]
				if not pd.isna(parsed):
					cell.value = parsed.to_pydatetime()
					cell.number_format = "dd/mm/yyyy"

		def _align_columns(column_indices: list[int], horizontal: str):
			for col_idx in column_indices:
				if col_idx > ws.max_column:
					continue
				for row_idx in range(2, ws.max_row + 1):
					cell = ws.cell(row=row_idx, column=col_idx)
					if cell.value is not None and str(cell.value).strip() != "":
						cell.alignment = Alignment(horizontal=horizontal)

		# Requested Step 3 layout formatting.
		_align_columns([2, 3, 4, 6, 9, 23, 26], "right")

		# Column AD should not show decimals.
		ad_col_idx = 30  # AD
		if ad_col_idx <= ws.max_column:
			for row_idx in range(2, ws.max_row + 1):
				cell = ws.cell(row=row_idx, column=ad_col_idx)
				if cell.value is None or str(cell.value).strip() == "":
					continue
				try:
					numeric_value = pd.to_numeric([cell.value], errors="coerce")[0]
					if not pd.isna(numeric_value):
						cell.value = int(float(numeric_value))
				except Exception:
					pass
				cell.number_format = "0"

		# Autofit all columns based on header/content length.
		for col_idx in range(1, ws.max_column + 1):
			max_len = 0
			for row_idx in range(1, ws.max_row + 1):
				value = ws.cell(row=row_idx, column=col_idx).value
				if value is None:
					continue
				value_len = len(str(value))
				if value_len > max_len:
					max_len = value_len
			ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

	output.seek(0)
	return output.getvalue()


def extract_mmddyy_token(file_name: str) -> str | None:
	"""[STEP 4] Extract 6-digit date token (MMDDYY) from file name for output naming."""
	name = Path(file_name).name
	match = re.search(r"(\d{6,8})", name)
	if not match:
		return None
	token = match.group(1)
	if len(token) == 8:
		return token[:4] + token[-2:]
	return token


def build_step3_call_type_file(source_df: pd.DataFrame) -> pd.DataFrame:
	"""[STEP 3] Insert Call Type column at position 0 with mapped values from Remark Type (Outgoing→Manual, Follow Up→Predictive)."""
	output_df = source_df.copy()

	remark_type_col = pick_column(
		list(output_df.columns),
		["Remark Type", "remark_type", "REMARK TYPE", "RemarkType"],
	)
	if remark_type_col is None:
		raise ValueError("Step 3 requires a 'Remark Type' column in the Call Logs file.")

	mapped_call_type = output_df[remark_type_col].map(map_step3_call_type)
	output_df.insert(0, "Call Type", mapped_call_type)
	return output_df


def to_polars_daily_call_logs_source(source_df: pd.DataFrame) -> pl.DataFrame:
	"""[STEP 2] Normalize pandas DataFrame object columns to pure strings before converting to Polars (fixes mixed type issues)."""
	prepared = source_df.copy()

	# Normalize object columns so Polars doesn't fail on mixed string/float cells.
	for col_name in prepared.columns:
		if prepared[col_name].dtype == "object":
			prepared[col_name] = prepared[col_name].map(lambda val: None if pd.isna(val) else str(val))

	return pl.from_pandas(prepared)


def update_feedback_workbook(
	feedback_file,
	merged_df: pd.DataFrame,
	feedback_account_col: str,
	feedback_cpm_col: str | None,
	feedback_account_id_col: str | None,
	feedback_contact_col: str | None,
	workbook_password: str | None = None,
) -> bytes:
	"""[FILL MISSING VALUES] Update feedback workbook in-place by writing only missing values."""
	workbook = _load_workbook_bytes(feedback_file, workbook_password=workbook_password)
	worksheet = workbook.active

	header_row = 1
	header_map = {}
	for cell in worksheet[header_row]:
		if cell.value is None:
			continue
		header_map[standardize_column_name(cell.value)] = cell.column

	account_col_idx = header_map.get(standardize_column_name(feedback_account_col))
	cpm_col_idx = header_map.get(standardize_column_name(feedback_cpm_col)) if feedback_cpm_col else None
	account_id_col_idx = header_map.get(standardize_column_name(feedback_account_id_col)) if feedback_account_id_col else None
	contact_col_idx = header_map.get(standardize_column_name(feedback_contact_col)) if feedback_contact_col else None

	if account_col_idx is None:
		raise ValueError("Could not find the account_number column in the feedback workbook header.")

	lookup_by_row = merged_df.set_index("_ROW_NUMBER", drop=False)
	for row_number, row in lookup_by_row.iterrows():
		excel_row = int(row_number)
		if cpm_col_idx and feedback_cpm_col and feedback_cpm_col in row:
			value = row.get(feedback_cpm_col)
			if not is_missing_like(value):
				worksheet.cell(row=excel_row, column=cpm_col_idx).value = value
		if account_id_col_idx and feedback_account_id_col and feedback_account_id_col in row:
			value = row.get(feedback_account_id_col)
			if not is_missing_like(value):
				worksheet.cell(row=excel_row, column=account_id_col_idx).value = value
		if contact_col_idx and feedback_contact_col and feedback_contact_col in row:
			value = row.get(feedback_contact_col)
			if not is_missing_like(value):
				worksheet.cell(row=excel_row, column=contact_col_idx).value = value

	output = BytesIO()
	workbook.save(output)
	output.seek(0)
	return output.getvalue()


DAILY_CALL_LOGS_SCHEMA = {
	"S.No": pl.Int64,
	"Date": pl.Utf8,
	"Time": pl.Utf8,
	"Debtor": pl.Utf8,
	"Account No.": pl.Int64,
	"Card No.": pl.Utf8,
	"Service No.": pl.Utf8,
	"DPD": pl.Int64,
	"Call Status": pl.Utf8,
	"Status": pl.Utf8,
	"Remark": pl.Utf8,
	"Remark By": pl.Utf8,
	"Remark Type": pl.Utf8,
	"Field Visit Date": pl.Utf8,
	"Collector": pl.Utf8,
	"Client": pl.Utf8,
	"Product Description": pl.Utf8,
	"Product Type": pl.Utf8,
	"Batch No": pl.Utf8,
	"Account Type": pl.Utf8,
	"Relation": pl.Utf8,
	"PTP Amount": pl.Utf8,
	"Next Call": pl.Utf8,
	"PTP Date": pl.Utf8,
	"Claim Paid Amount": pl.Utf8,
	"Claim Paid Date": pl.Utf8,
	"Dialed Number": pl.Utf8,
	"Days Past Write Off": pl.Utf8,
	"Balance": pl.Utf8,
	"Contact Type": pl.Utf8,
	"Cycle": pl.Utf8,
	"Call Duration": pl.Utf8,
	"Talk Time Duration": pl.Utf8,
}


LOXON_DAILY_REMARK_SCHEMA = {
	"S.No": pl.Int64,
	"Date": pl.Date,
	"Time": pl.Datetime,
	"Debtor": pl.Utf8,
	"Account No.": pl.Int64,
	"Card No.": pl.Utf8,
	"Service No.": pl.Utf8,
	"DPD": pl.Int64,
	"Reason For Default": pl.Utf8,
	"Call Status": pl.Utf8,
	"Status": pl.Utf8,
	"Remark": pl.Utf8,
	"Remark By": pl.Utf8,
	"Remark Type": pl.Utf8,
	"Field Visit Date": pl.Utf8,
	"Collector": pl.Utf8,
	"Client": pl.Utf8,
	"Product Description": pl.Utf8,
	"Product Type": pl.Utf8,
	"Batch No": pl.Utf8,
	"Account Type": pl.Utf8,
	"Relation": pl.Utf8,
	"PTP Amount": pl.Float64,
	"Next Call": pl.Utf8,
	"PTP Date": pl.Utf8,
	"Claim Paid Amount": pl.Float64,
	"Claim Paid Date": pl.Utf8,
	"Dialed Number": pl.Utf8,
	"Days Past Write Off": pl.Int64,
	"Balance": pl.Float64,
	"Contact Type": pl.Utf8,
	"Call Duration": pl.Int64,
	"Talk Time Duration": pl.Int64,
}


def load_merge_account_mappings() -> tuple[dict, dict]:
	"""[STEP 4] Load agent code mapping dictionaries from JSON resources (BCRM VOLARE → fullname mappings)."""
	with open("./resources/agent_code_bcrm_volare.json", "r") as file:
		agent_code_bcrm_volare = json.load(file)
	with open("./resources/agent_code_volare_fullname.json", "r") as file:
		agent_code_volare_fullname = json.load(file)
	return agent_code_bcrm_volare, agent_code_volare_fullname


def load_loxon_references() -> tuple[pl.DataFrame, pl.DataFrame]:
	"""[STEPS 2, 5] Load reference mapping DataFrames from Maya reference workbook (STATUS and AGENT sheets)."""
	status_ref = pl.read_excel("./resources/maya_reference.xlsx", sheet_name="STATUS")
	agent_ref = pl.read_excel("./resources/maya_reference.xlsx", sheet_name="AGENT")
	return status_ref, agent_ref


def time_to_seconds(time_str: str) -> int:
	"""[STEP 2, 5] Convert HH:MM:SS time string to total seconds integer."""
	if time_str is None:
		return 0
	if isinstance(time_str, int):
		return int(time_str)
	if isinstance(time_str, float):
		return int(time_str)
	hours, minutes, seconds = map(int, str(time_str).split(":"))
	return hours * 3600 + minutes * 60 + seconds


def add_seconds(start_time: datetime, seconds: int) -> datetime:
	"""[STEP 5] Add seconds to a datetime object to calculate call end time (used in LOXON upload)."""
	if start_time is None:
		return None
	return start_time + timedelta(seconds=int(seconds or 0))


def seconds_to_time(seconds: int) -> time:
	"""[STEP 5] Convert total seconds to time(HH:MM:SS) object for LOXON upload duration field."""
	hours = (int(seconds) // 3600) % 24
	minutes = (int(seconds) % 3600) // 60
	secs = int(seconds) % 60
	return time(hour=hours, minute=minutes, second=secs)


def extract_rfd(text: str) -> str | None:
	"""[STEP 5] Extract Reason For Delay (RFD) from remark text using regex pattern 'RFD: (.*?) |' (used in LOXON upload)."""
	if text is None:
		return None
	match = re.search(r"RFD: (.*?) \|", str(text))
	return match.group(1) if match else None


def product_name(input_value) -> str:
	"""[STEP 5] Map account number prefix to product name (6=MayaCredit, 4=negosyoAdvance, 9=MAYA_FLEXI_ENTERPRISE_LOAN)."""
	input_string = str(input_value)
	if input_string.startswith("6"):
		return "MayaCredit"
	if input_string.startswith("4"):
		return "negosyoAdvance"
	if input_string.startswith("9"):
		return "MAYA_FLEXI_ENTERPRISE_LOAN"
	return ""


def format_phone_number(phone) -> str:
	"""[STEP 5] Normalize phone numbers to 63xxxxxxxxxxx format (remove leading 0, add 63 prefix, strip +)."""
	phone = str(phone)
	if phone.startswith("09"):
		return "63" + phone[1:]
	if phone.startswith("9"):
		return "63" + phone
	if phone.startswith("+63"):
		return phone[1:]
	return phone


def build_daily_call_logs(tx_status: pl.DataFrame, dispo_list: pl.DataFrame) -> tuple[pl.DataFrame, pl.DataFrame]:
	"""[STEP 2] Transform daily remarks into call logs with status mapping, date parsing, and duration conversions."""
	drop_cols = [
		"Black Case No.", "Red Case No.", "Court Name", "Lawyer", "Legal Stage", "Legal Status",
		"Next Legal Follow up", "Old IC", "I.C Issue Date", "Bank Code", "Over Limit Amount",
		"Min Payment", "Due Date", "Monthly Installment", "30 Days", "MIA", "Area", "Debtor ID",
	]
	existing_drop_cols = [col for col in drop_cols if col in tx_status.columns]
	if existing_drop_cols:
		tx_status = tx_status.drop(existing_drop_cols)

	tx_status = tx_status.with_columns(
		pl.col("PTP Amount").cast(pl.Utf8).str.replace_all(",", "").cast(pl.Float64, strict=False).alias("PTP Amount"),
		pl.col("Claim Paid Amount").cast(pl.Utf8).str.replace_all(",", "").cast(pl.Float64, strict=False).alias("Claim Paid Amount"),
		pl.col("Balance").cast(pl.Utf8).str.replace_all(",", "").cast(pl.Float64, strict=False).alias("Balance"),
		pl.coalesce([
			pl.col("Date").cast(pl.Date, strict=False),
			pl.col("Date").cast(pl.Utf8).str.strptime(pl.Date, "%Y-%m-%d", strict=False),
			pl.col("Date").cast(pl.Utf8).str.strptime(pl.Date, "%m/%d/%Y", strict=False),
			pl.col("Date").cast(pl.Utf8).str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S", strict=False).dt.date(),
			pl.col("Date").cast(pl.Utf8).str.strptime(pl.Datetime, "%m/%d/%Y %H:%M:%S", strict=False).dt.date(),
		]).alias("Date"),
		pl.coalesce([
			pl.col("Time").cast(pl.Time, strict=False),
			pl.col("Time").cast(pl.Utf8).str.strptime(pl.Time, "%I:%M:%S %p", strict=False),
			pl.col("Time").cast(pl.Utf8).str.strptime(pl.Time, "%H:%M:%S", strict=False),
			pl.col("Time").cast(pl.Utf8).str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S", strict=False).dt.time(),
			pl.col("Time").cast(pl.Utf8).str.strptime(pl.Datetime, "%m/%d/%Y %H:%M:%S", strict=False).dt.time(),
			pl.col("Time").cast(pl.Utf8).str.strptime(pl.Datetime, "%m/%d/%Y %I:%M:%S %p", strict=False).dt.time(),
		]).alias("Time"),
		pl.lit(None).alias("Reason For Default"),
		pl.col("Call Duration").map_elements(time_to_seconds, return_dtype=pl.Int64).alias("Call Duration"),
		pl.col("Talk Time Duration").map_elements(time_to_seconds, return_dtype=pl.Int64).alias("Talk Time Duration"),
	)

	tx_status = tx_status.with_columns(
		(pl.col("Date").dt.date().cast(pl.Datetime) + pl.col("Time").cast(pl.Duration)).alias("Time")
	)

	exclude_status = [None, "ABORT", "BP", "NEW", "REACTIVE", "FS", "PP"]
	tx_status = tx_status.filter(~(pl.col("Status").is_in(exclude_status)))
	tx_status = tx_status.with_columns(pl.arange(1, tx_status.height + 1).alias("S.No"))

	volare_dispo = dict(zip(dispo_list["VOLARE STATUS"].to_list(), dispo_list["PROPOSED DISPOSITION"].to_list()))
	call_logs = tx_status.with_columns(
		pl.col("Status").replace_strict(volare_dispo, default=None).alias("Status")
	).filter(pl.col("Status").is_not_null())

	daily_remark_columns = [
		"S.No", "Date", "Time", "Debtor", "Account No.", "Cycle", "Card No.", "Service No.", "DPD",
		"Reason For Default", "Call Status", "Status", "Remark", "Remark By", "Remark Type", "Field Visit Date",
		"Collector", "Client", "Product Description", "Product Type", "Batch No", "Account Type", "Relation",
		"PTP Amount", "Next Call", "PTP Date", "Claim Paid Amount", "Claim Paid Date", "Dialed Number",
		"Days Past Write Off", "Balance", "Contact Type", "Call Duration", "Talk Time Duration",
	]
	for extra_col in ["Hierarchy", "result", "channel"]:
		if extra_col in tx_status.columns:
			daily_remark_columns.append(extra_col)
	daily_remark = tx_status.select(daily_remark_columns)
	return daily_remark, call_logs


def build_merged_accounts(
	endorsement_file,
	passkey: str,
	start_date: date,
	end_date: date,
	agent_code_bcrm_volare: dict,
	agent_code_volare_fullname: dict,
) -> pl.DataFrame:
	"""[STEP 4] Read encrypted endorsement file, merge ACTIVE+POUT sheets, apply agent mappings, filter by date range."""
	file_bytes = _file_obj_to_bytes(endorsement_file)
	decrypted_bytes = _decrypt_excel_bytes(file_bytes, passkey)

	active_sheet = pl.read_excel(
		BytesIO(decrypted_bytes),
		sheet_name="ACTIVE",
		schema_overrides={"ACCOUNT_NUMBER_LAST_SET_TO_ARREARS_DATE": pl.Utf8},
	)
	pout_sheet = pl.read_excel(
		BytesIO(decrypted_bytes),
		sheet_name="POUT",
		schema_overrides={"ACCOUNT_NUMBER_LAST_SET_TO_ARREARS_DATE": pl.Utf8},
	)

	active_sheet = active_sheet.with_columns(
		pl.lit("ACTIVE").alias("REMARKS"),
		pl.lit(None).alias("PULLED OUT DATE"),
	)

	base_cols = [
		"PULLED OUT DATE", "REMARKS", "PLACEMENT", "ACCOUNT NUMBER", "ENDO STAT", "CHCODE", "TAGGING",
		"DPD BUCKET", "DPD_", "MOBILE PROPER", "OB", "FRESH/SPILLOVER", "RECEIVED DATE", "AS_OF", "CPM_ID",
		"NAME", "FIRST_NAME", "LAST_NAME", "BIRTH_DATE", "ACCOUNT_ID", "PRODUCT_NAME",
	]
	active_sheet = active_sheet.select([col for col in base_cols if col in active_sheet.columns])

	pout_sheet = pout_sheet.filter((pl.col("PULLED OUT DATE") >= start_date) & (pl.col("PULLED OUT DATE") <= end_date))
	pout_sheet = pout_sheet.select([col for col in base_cols if col in pout_sheet.columns])

	column_types = {
		"PULLED OUT DATE": pl.Date,
		"REMARKS": pl.Utf8,
		"PLACEMENT": pl.Utf8,
		"ACCOUNT NUMBER": pl.Int64,
		"ENDO STAT": pl.Utf8,
		"CHCODE": pl.Utf8,
		"TAGGING": pl.Utf8,
		"DPD BUCKET": pl.Utf8,
		"DPD_": pl.Int64,
		"MOBILE PROPER": pl.Utf8,
		"OB": pl.Float64,
		"FRESH/SPILLOVER": pl.Utf8,
		"RECEIVED DATE": pl.Date,
		"AS_OF": pl.Date,
		"CPM_ID": pl.Utf8,
		"NAME": pl.Utf8,
		"FIRST_NAME": pl.Utf8,
		"LAST_NAME": pl.Utf8,
		"BIRTH_DATE": pl.Date,
		"ACCOUNT_ID": pl.Utf8,
		"PRODUCT_NAME": pl.Utf8,
	}

	active_sheet = cast_columns(active_sheet, column_types)
	pout_sheet = cast_columns(pout_sheet, column_types)

	merged_accounts = active_sheet.vstack(pout_sheet)
	merged_accounts = merged_accounts.with_columns(
		pl.col("TAGGING").map_elements(lambda x: agent_code_bcrm_volare.get(x, None), return_dtype=pl.Utf8).alias("VOLARE TAGGING")
	)
	merged_accounts = merged_accounts.with_columns(
		pl.col("VOLARE TAGGING").map_elements(lambda x: agent_code_volare_fullname.get(x, None), return_dtype=pl.Utf8).alias("AGENT NAME")
	)

	merged_accounts = merged_accounts.select([
		"PULLED OUT DATE", "REMARKS", "PLACEMENT", "ACCOUNT NUMBER", "ENDO STAT", "CHCODE", "TAGGING", "VOLARE TAGGING",
		"AGENT NAME", "DPD BUCKET", "DPD_", "MOBILE PROPER", "OB", "FRESH/SPILLOVER", "RECEIVED DATE", "AS_OF", "CPM_ID",
		"NAME", "FIRST_NAME", "LAST_NAME", "BIRTH_DATE", "ACCOUNT_ID", "PRODUCT_NAME",
	])

	return merged_accounts.filter(pl.col("AS_OF").is_not_null()).sort("AS_OF", descending=True)


def build_loxon_upload(
	merged_accounts: pl.DataFrame,
	daily_remark: pl.DataFrame,
	status_ref: pl.DataFrame,
	agent_ref: pl.DataFrame,
) -> pl.DataFrame:
	"""[DEPRECATED] Legacy builder function (Step 5 now uses exact LOXON Upload.py logic directly in form submission)."""
	if "Hierarchy" not in daily_remark.columns:
		daily_remark = daily_remark.with_columns(pl.lit(0).alias("Hierarchy"))
	if "result" not in daily_remark.columns:
		daily_remark = daily_remark.with_columns(pl.lit(None).alias("result"))
	if "channel" not in daily_remark.columns:
		daily_remark = daily_remark.with_columns(pl.lit(None).alias("channel"))

	daily_remark = daily_remark.filter(
		~(
			pl.col("Remark").str.starts_with("System Auto Update")
			| pl.col("Remark").str.starts_with("Updates when case reassign")
		)
	)
	daily_remark = daily_remark.with_columns(pl.col("Call Duration").fill_null(0).alias("Call Duration"))

	daily_remark = daily_remark.join(status_ref, left_on="Status", right_on="VOLARE STATUS", how="left")
	daily_remark = daily_remark.join(
		agent_ref.select(["VOLARE USERNAME", "AGENT NAME"]),
		left_on="Remark By",
		right_on="VOLARE USERNAME",
		how="left",
	)

	daily_remark = daily_remark.with_columns(
		pl.struct(["Time", "Call Duration"]).map_elements(
			lambda x: add_seconds(x["Time"], x["Call Duration"]),
			return_dtype=pl.Datetime,
		).alias("END"),
		pl.col("Remark").map_elements(extract_rfd, return_dtype=pl.Utf8).alias("RFD"),
		pl.col("Call Duration").map_elements(seconds_to_time, return_dtype=pl.Time).alias("duration"),
		pl.col("Account No.").map_elements(product_name, return_dtype=pl.Utf8).alias("product_name"),
	)

	daily_remark = daily_remark.with_columns(pl.col("Hierarchy").fill_null(0)).sort("Hierarchy", descending=True)
	daily_remark_unique = daily_remark.unique(subset="Account No.", keep="first", maintain_order=True)

	loxon_upload = daily_remark_unique.select([
		"Time", "result", "Remark", "Account No.", "Dialed Number", "PTP Date", "PTP Amount", "RFD", "END",
		"duration", "AGENT NAME", "product_name", "channel",
	])
	loxon_upload = loxon_upload.with_columns(
		pl.lit(None).alias("outsource_case_id"),
		pl.lit("Madrid").alias("outsource_partner_alias"),
		pl.lit(None).alias("skip_phone_number"),
		pl.lit(None).alias("skip_email"),
		pl.lit(None).alias("nonvoice_template"),
		pl.col("Time").alias("event_datetime_pht"),
		pl.col("Time").alias("datalate_processed_ts_pht"),
		pl.col("Dialed Number").map_elements(format_phone_number, return_dtype=pl.Utf8).alias("Dialed Number"),
	)

	loxon_upload = loxon_upload.join(
		merged_accounts.select(["ACCOUNT NUMBER", "AGENT NAME", "CPM_ID", "ACCOUNT_ID", "MOBILE PROPER"]),
		left_on="Account No.",
		right_on="ACCOUNT NUMBER",
		how="left",
	)

	loxon_upload = loxon_upload.rename(
		{
			"Time": "call_start",
			"Remark": "comment",
			"Account No.": "account_number",
			"Dialed Number": "number_contacted",
			"PTP Date": "ptp_date",
			"PTP Amount": "ptp_amount",
			"RFD": "reason_for_delay",
			"END": "call_end",
			"AGENT NAME": "collector_full_name",
			"ACCOUNT_ID": "account_id",
			"CPM_ID": "cpm_id",
			"channel": "communication_channel",
		}
	)

	loxon_upload = loxon_upload.with_columns(
		pl.when((pl.col("number_contacted").is_null()) & (pl.col("communication_channel") == "VOICE"))
		.then(pl.col("MOBILE PROPER"))
		.otherwise(pl.col("number_contacted"))
		.alias("number_contacted"),
		pl.when(pl.col("collector_full_name").is_null())
		.then(pl.col("AGENT NAME_right"))
		.otherwise(pl.col("collector_full_name"))
		.alias("collector_full_name"),
		pl.col("ptp_amount").replace(0.0, None),
	)

	loxon_upload = loxon_upload.with_columns(
		pl.col("number_contacted").map_elements(format_phone_number, return_dtype=pl.Utf8).alias("number_contacted")
	)

	return loxon_upload.select([
		"outsource_case_id", "outsource_partner_alias", "event_datetime_pht", "result", "comment", "cpm_id", "account_id",
		"account_number", "product_name", "communication_channel", "number_contacted", "ptp_date", "ptp_amount", "reason_for_delay",
		"call_start", "call_end", "duration", "collector_full_name", "skip_phone_number", "skip_email", "nonvoice_template",
		"datalate_processed_ts_pht",
	])


def first_non_blank(series: pd.Series):
	"""[FILL MISSING VALUES] Return first non-blank value from a pandas Series (used in aggregation for POUT lookup)."""
	for value in series:
		if is_missing_like(value):
			continue
		return value
	return None


def fill_missing_values_pipeline(
	feedback_file,
	endorsement_source,
	workbook_password: str | None = None,
) -> tuple[bytes, pd.DataFrame, dict[str, int], str]:
	"""[FILL MISSING VALUES] Fill CPM ID, Account ID, and Number Contacted from endorsement data."""
	password = workbook_password or DEFAULT_WORKBOOK_PASSWORD
	feedback_df = read_excel(feedback_file, workbook_password=password)
	endorsement_df = read_endorsement_pout(endorsement_source, workbook_password=password)
	endorsement_active_df = read_endorsement_active(endorsement_source, workbook_password=password)

	feedback_std_map = {col: standardize_column_name(col) for col in feedback_df.columns}
	endorsement_std_map = {col: standardize_column_name(col) for col in endorsement_df.columns}
	active_std_map = {col: standardize_column_name(col) for col in endorsement_active_df.columns}

	feedback_std_to_raw = {v: k for k, v in feedback_std_map.items()}
	endorsement_std_to_raw = {v: k for k, v in endorsement_std_map.items()}
	active_std_to_raw = {v: k for k, v in active_std_map.items()}

	feedback_account_std = pick_column(list(feedback_std_to_raw.keys()), ["account_number", "account_no", "account", "account_num"])
	feedback_cpm_std = pick_column(list(feedback_std_to_raw.keys()), ["cpm_id", "cpm"])
	feedback_account_id_std = pick_column(list(feedback_std_to_raw.keys()), ["account_id", "acct_id"])
	feedback_contact_std = pick_column(list(feedback_std_to_raw.keys()), ["number_contacted", "contact_number", "mobile_number", "phone_number"])

	endorsement_account_std = pick_column(list(endorsement_std_to_raw.keys()), ["account_number", "account_no", "account", "account_num"])
	endorsement_cpm_std = pick_column(list(endorsement_std_to_raw.keys()), ["cpm_id", "cpm"])
	endorsement_account_id_std = pick_column(list(endorsement_std_to_raw.keys()), ["account_id", "acct_id"])
	endorsement_mobile_std = pick_column(list(endorsement_std_to_raw.keys()), ["mobile_proper", "mobile", "phone_number", "contact_number"])

	if feedback_account_std is None or endorsement_account_std is None:
		raise ValueError("Missing account_number/account column in one of the files.")
	if feedback_cpm_std is None and feedback_account_id_std is None:
		raise ValueError("Feedback file must have cpm_id and/or account_id columns to fill.")
	if endorsement_cpm_std is None and endorsement_account_id_std is None:
		raise ValueError("Endorsement file must have cpm_id and/or account_id columns as source values.")

	feedback_account_col = feedback_std_to_raw[feedback_account_std]
	endorsement_account_col = endorsement_std_to_raw[endorsement_account_std]
	feedback_contact_col = feedback_std_to_raw[feedback_contact_std] if feedback_contact_std is not None else None
	feedback_cpm_col = feedback_std_to_raw[feedback_cpm_std] if feedback_cpm_std is not None else None
	feedback_account_id_col = feedback_std_to_raw[feedback_account_id_std] if feedback_account_id_std is not None else None

	working_feedback = feedback_df.copy()
	working_feedback["_ROW_NUMBER"] = working_feedback.index + 2
	working_feedback["_ACCOUNT_KEY"] = working_feedback[feedback_account_col].map(to_account_key)
	working_feedback = working_feedback[working_feedback["_ACCOUNT_KEY"] != ""].copy()

	working_endorsement = endorsement_df.copy()
	working_endorsement["_ACCOUNT_KEY"] = working_endorsement[endorsement_account_col].map(to_account_key)
	working_endorsement = working_endorsement[working_endorsement["_ACCOUNT_KEY"] != ""].copy()

	lookup_cols = ["_ACCOUNT_KEY"]
	if endorsement_cpm_std is not None:
		lookup_cols.append(endorsement_std_to_raw[endorsement_cpm_std])
	if endorsement_account_id_std is not None:
		lookup_cols.append(endorsement_std_to_raw[endorsement_account_id_std])
	pout_mobile_col = endorsement_std_to_raw[endorsement_mobile_std] if endorsement_mobile_std is not None else None
	if pout_mobile_col is not None:
		lookup_cols.append(pout_mobile_col)

	lookup_df = working_endorsement[lookup_cols].copy()
	agg_map = {col: first_non_blank for col in lookup_cols if col != "_ACCOUNT_KEY"}
	lookup_df = lookup_df.groupby("_ACCOUNT_KEY", as_index=False).agg(agg_map)
	if pout_mobile_col is not None and pout_mobile_col in lookup_df.columns:
		lookup_df = lookup_df.rename(columns={pout_mobile_col: "MOBILE_PROPER_POUT"})

	active_mobile_std = pick_column(list(active_std_to_raw.keys()), ["mobile_proper", "mobile", "phone_number", "contact_number"])
	active_account_std = pick_column(list(active_std_to_raw.keys()), ["account_number", "account_no", "account", "account_num"])
	if active_account_std is not None and active_mobile_std is not None:
		active_account_col = active_std_to_raw[active_account_std]
		active_mobile_col = active_std_to_raw[active_mobile_std]
		working_active = endorsement_active_df.copy()
		working_active["_ACCOUNT_KEY"] = working_active[active_account_col].map(to_account_key)
		working_active = working_active[working_active["_ACCOUNT_KEY"] != ""].copy()
		active_lookup_df = working_active[["_ACCOUNT_KEY", active_mobile_col]].copy()
		active_lookup_df = active_lookup_df.groupby("_ACCOUNT_KEY", as_index=False).agg({active_mobile_col: first_non_blank})
		active_lookup_df = active_lookup_df.rename(columns={active_mobile_col: "MOBILE_PROPER_ACTIVE"})
	else:
		active_lookup_df = pd.DataFrame(columns=["_ACCOUNT_KEY", "MOBILE_PROPER_ACTIVE"])

	merged_df = working_feedback.merge(lookup_df, on="_ACCOUNT_KEY", how="left")
	merged_df = merged_df.merge(active_lookup_df, on="_ACCOUNT_KEY", how="left")

	stats = {"cpm_filled": 0, "account_id_filled": 0, "contact_filled": 0, "normalized_contacts": 0}

	if feedback_cpm_col is not None and endorsement_cpm_std is not None:
		source_col = endorsement_std_to_raw[endorsement_cpm_std]
		source_col = f"{source_col}_x" if f"{source_col}_x" in merged_df.columns else source_col
		missing_mask = merged_df[feedback_cpm_col].map(is_missing_like)
		source_has_value = ~merged_df[source_col].map(is_missing_like) if source_col in merged_df.columns else pd.Series([False] * len(merged_df))
		fill_mask = missing_mask & source_has_value
		stats["cpm_filled"] = int(fill_mask.sum())
		merged_df.loc[fill_mask, feedback_cpm_col] = merged_df.loc[fill_mask, source_col]

	if feedback_account_id_col is not None and endorsement_account_id_std is not None:
		source_col = endorsement_std_to_raw[endorsement_account_id_std]
		source_col = f"{source_col}_x" if f"{source_col}_x" in merged_df.columns else source_col
		missing_mask = merged_df[feedback_account_id_col].map(is_missing_like)
		source_has_value = ~merged_df[source_col].map(is_missing_like) if source_col in merged_df.columns else pd.Series([False] * len(merged_df))
		fill_mask = missing_mask & source_has_value
		stats["account_id_filled"] = int(fill_mask.sum())
		merged_df.loc[fill_mask, feedback_account_id_col] = merged_df.loc[fill_mask, source_col]

	if feedback_contact_col is not None:
		if "MOBILE_PROPER_ACTIVE" in merged_df.columns:
			missing_mask = merged_df[feedback_contact_col].map(is_missing_like)
			source_has_value = ~merged_df["MOBILE_PROPER_ACTIVE"].map(is_missing_like)
			fill_mask = missing_mask & source_has_value
			stats["contact_filled"] += int(fill_mask.sum())
			merged_df.loc[fill_mask, feedback_contact_col] = merged_df.loc[fill_mask, "MOBILE_PROPER_ACTIVE"]
		if "MOBILE_PROPER_POUT" in merged_df.columns:
			missing_mask = merged_df[feedback_contact_col].map(is_missing_like)
			source_has_value = ~merged_df["MOBILE_PROPER_POUT"].map(is_missing_like)
			fill_mask = missing_mask & source_has_value
			stats["contact_filled"] += int(fill_mask.sum())
			merged_df.loc[fill_mask, feedback_contact_col] = merged_df.loc[fill_mask, "MOBILE_PROPER_POUT"]
		before = merged_df[feedback_contact_col].copy()
		merged_df[feedback_contact_col] = merged_df[feedback_contact_col].map(normalize_number_contacted)
		stats["normalized_contacts"] = int((before.astype(str) != merged_df[feedback_contact_col].astype(str)).sum())

	unresolved_rows = []
	for _, row in merged_df.iterrows():
		missing_cpm = feedback_cpm_col is not None and is_missing_like(row.get(feedback_cpm_col))
		missing_account_id = feedback_account_id_col is not None and is_missing_like(row.get(feedback_account_id_col))
		if not (missing_cpm or missing_account_id):
			continue
		account_value = row.get(feedback_account_col)
		account_key = to_account_key(account_value)
		lookup_present = False
		if endorsement_cpm_std is not None:
			source_col = endorsement_std_to_raw[endorsement_cpm_std]
			lookup_present = lookup_present or not is_missing_like(row.get(source_col)) or not is_missing_like(row.get(f"{source_col}_x"))
		if endorsement_account_id_std is not None:
			source_col = endorsement_std_to_raw[endorsement_account_id_std]
			lookup_present = lookup_present or not is_missing_like(row.get(source_col)) or not is_missing_like(row.get(f"{source_col}_x"))
		if not account_key:
			reason = "Missing account_number in feedback"
		elif not lookup_present:
			reason = "No matching endorsement record found for this account_number"
		else:
			reason = "Endorsement match found but source values were blank"
		unresolved_rows.append({"Feedback Row": int(row.get("_ROW_NUMBER", 0) or 0), "account_number": account_value, "Reason": reason})

	drop_cols = [col for col in merged_df.columns if col.endswith("_x") or col.endswith("_y")]
	for extra_col in ["_ACCOUNT_KEY", "_ROW_NUMBER", "MOBILE_PROPER_ACTIVE", "MOBILE_PROPER_POUT"]:
		if extra_col in merged_df.columns:
			drop_cols.append(extra_col)
	export_df = merged_df.drop(columns=drop_cols, errors="ignore")
	output_name = f"{Path(feedback_file.name).stem}.xlsx"
	if Path(feedback_file.name).suffix.lower() == ".csv":
		output_bytes = to_output_excel(export_df)
	else:
		output_bytes = update_feedback_workbook(
			feedback_file,
			merged_df,
			feedback_account_col=feedback_account_col,
			feedback_cpm_col=feedback_cpm_col,
			feedback_account_id_col=feedback_account_id_col,
			feedback_contact_col=feedback_contact_col,
			workbook_password=password,
		)

	return output_bytes, pd.DataFrame(unresolved_rows), stats, output_name


def run_ui():
	"""[MAIN] Single-page UI with one button to run the full LOXON pipeline end-to-end."""
	st.markdown(
		"""
		<style>
			.block-container {
				max-width: 940px;
				padding-top: 1.4rem;
				padding-bottom: 2rem;
			}
			div[data-testid="stProgressBar"] > div > div {
				transition: width 0.35s ease;
			}
			.pipeline-card {
				border: 1px solid rgba(49, 51, 63, 0.16);
				border-radius: 12px;
				padding: 1rem;
				background: rgba(255, 255, 255, 0.86);
			}
		</style>
		""",
		unsafe_allow_html=True,
	)

	state_defaults = {
		"pipeline_running": False,
		"pipeline_logs": [],
		"pipeline_percent": 0,
		"pipeline_status": "Waiting to start...",
		"step1_combined_daily_remark_bytes": None,
		"step1_combined_daily_remark_name": None,
		"step2_daily_remark_bytes": None,
		"step2_daily_remark_name": None,
		"step2_call_logs_bytes": None,
		"step2_call_logs_name": None,
		"step3_call_type_bytes": None,
		"step3_call_type_name": None,
		"step4_merged_accounts_bytes": None,
		"step4_merged_accounts_name": None,
		"step5_loxon_upload_bytes": None,
		"step5_loxon_upload_name": None,
		"final_output_bytes": None,
		"final_output_name": None,
		"final_unresolved_df": None,
		"final_stats": None,
	}
	for key, default_value in state_defaults.items():
		if key not in st.session_state:
			st.session_state[key] = default_value

	def clear_pipeline_outputs():
		for key in [
			"step1_combined_daily_remark_bytes",
			"step1_combined_daily_remark_name",
			"step2_daily_remark_bytes",
			"step2_daily_remark_name",
			"step2_call_logs_bytes",
			"step2_call_logs_name",
			"step3_call_type_bytes",
			"step3_call_type_name",
			"step4_merged_accounts_bytes",
			"step4_merged_accounts_name",
			"step5_loxon_upload_bytes",
			"step5_loxon_upload_name",
			"final_output_bytes",
			"final_output_name",
			"final_unresolved_df",
			"final_stats",
		]:
			st.session_state[key] = None

	def append_status(percent: int, step_message: str, detail: str | None = None):
		st.session_state.pipeline_percent = max(0, min(100, percent))
		st.session_state.pipeline_status = step_message
		if detail:
			st.session_state.pipeline_logs.append(f"{percent}% - {detail}")
		else:
			st.session_state.pipeline_logs.append(f"{percent}% - {step_message}")
		st.session_state.pipeline_logs = st.session_state.pipeline_logs[-14:]

	def rewind_if_possible(file_obj):
		if hasattr(file_obj, "seek"):
			file_obj.seek(0)

	st.title("LOXON Remarks Auto")
	st.caption("Single streamlined workflow with one action button.")

	disabled_inputs = st.session_state.pipeline_running

	left, center, right = st.columns([1, 4, 1])
	with center:
		st.markdown("<div class='pipeline-card'>", unsafe_allow_html=True)
		st.subheader("Pipeline Input")

		uploaded_files = st.file_uploader(
			"Upload Required Files (exactly 3)",
			type=["xlsx", "xls", "xlsb", "csv"],
			accept_multiple_files=True,
			key="single_required_files",
			disabled=disabled_inputs,
		)

		st.caption(f"Fixed Server: {DEFAULT_SERVER_MASTERFILE_DIR}")
		endo_masterfile_name = st.text_input(
			"Endo Selection (file name or relative path in MASTERFILE)",
			value=DEFAULT_SERVER_ENDORSEMENT_FILE,
			help="Example: MAYA ENDORSEMENT 04152026.xlsx or MARCH 2026\\MAYA ENDORSEMENT 04152026.xlsx",
			key="single_endo_masterfile_name",
			disabled=disabled_inputs,
		)
		optional_endo_upload = st.file_uploader(
			"Optional: Upload Endo/Masterfile (overrides server file)",
			type=["xlsx"],
			accept_multiple_files=False,
			key="single_optional_endo_upload",
			disabled=disabled_inputs,
		)

		workbook_password = st.text_input(
			"Password",
			value=DEFAULT_WORKBOOK_PASSWORD,
			type="password",
			key="single_password",
			disabled=disabled_inputs,
		)

		date_cols = st.columns(2)
		with date_cols[0]:
			now_dt = datetime.now()
			start_date = st.date_input(
				"Start Date",
				value=datetime(now_dt.year, now_dt.month, 1),
				key="single_start_date",
				disabled=disabled_inputs,
			)
		with date_cols[1]:
			end_date = st.date_input(
				"End Date",
				value="today",
				key="single_end_date",
				disabled=disabled_inputs,
			)

		files_ready = len(uploaded_files or []) == 3
		endo_ready = bool(str(endo_masterfile_name).strip()) or optional_endo_upload is not None
		password_ready = bool(str(workbook_password).strip())
		can_start = files_ready and endo_ready and password_ready and not disabled_inputs

		start_process = st.button(
			"Start Process",
			type="primary",
			use_container_width=True,
			disabled=not can_start,
		)

		if not can_start:
			st.caption("Upload exactly 3 required files, set Endo selection (or upload optional Endo), and enter password to enable Start Process.")

		st.markdown("</div>", unsafe_allow_html=True)

	progress_bar = st.progress(st.session_state.pipeline_percent)
	status_slot = st.empty()
	status_slot.markdown(f"**{st.session_state.pipeline_percent}%** {st.session_state.pipeline_status}")
	st.text_area(
		"Status Feed",
		"\n".join(st.session_state.pipeline_logs),
		height=170,
		disabled=True,
		key="single_status_feed",
	)

	if start_process:
		st.session_state.pipeline_running = True
		st.session_state.pipeline_logs = []
		clear_pipeline_outputs()

		try:
			def set_progress(percent: int, step_message: str, detail: str):
				append_status(percent, step_message, detail)
				progress_bar.progress(st.session_state.pipeline_percent)
				status_slot.markdown(f"**{st.session_state.pipeline_percent}%** {st.session_state.pipeline_status}")

			set_progress(5, "Validating inputs...", "Checking required files, Endo source, and password.")

			uploaded_files_safe = uploaded_files or []
			if len(uploaded_files_safe) != 3:
				raise ValueError("Please upload exactly 3 required files.")

			set_progress(10, "Validating inputs...", "Resolving Endo source from optional upload or fixed server path.")
			selected_endo_source = optional_endo_upload
			selected_endo_name = optional_endo_upload.name if optional_endo_upload is not None else None
			if selected_endo_source is None:
				selected_endo_source, server_error = resolve_server_endorsement_file(endo_masterfile_name)
				if server_error:
					raise ValueError(server_error)
				selected_endo_name = selected_endo_source.name

			set_progress(15, "Uploading files...", "Preparing required files for dataset combination.")

			combine_file_1, combine_file_2, combine_file_3 = uploaded_files_safe
			rewind_if_possible(combine_file_1)
			rewind_if_possible(combine_file_2)
			rewind_if_possible(combine_file_3)
			set_progress(20, "Uploading files...", "Combining the 3 required files by header alignment.")
			combined_df = combine_three_files_by_header(combine_file_1, combine_file_2, combine_file_3, workbook_password=None)
			combined_date_token = resolve_report_date_token(combined_df)
			step1_date_token = combined_date_token[:4] + combined_date_token[-2:] if len(combined_date_token) == 8 else combined_date_token
			st.session_state.step1_combined_daily_remark_bytes = to_output_excel_preserve(
				combined_df,
				sheet_name="Daily Remark Report",
			)
			st.session_state.step1_combined_daily_remark_name = f"Daily_Remark_Report_{step1_date_token}.xlsx"

			set_progress(28, "Extracting columns...", "Preparing source frame for Daily Remark and Call Logs extraction.")

			tx_status = to_polars_daily_call_logs_source(combined_df.copy())
			dispo_list = pl.read_csv("./resources/maya_dispositions.csv")
			set_progress(34, "Extracting columns...", "Transforming and filtering statuses to build output datasets.")
			daily_remark, call_logs = build_daily_call_logs(tx_status, dispo_list)
			daily_remark_formats = {
				pl.Int8: "0",
				pl.Int16: "0",
				pl.Int32: "0",
				pl.Int64: "0",
				pl.Float32: "0.00",
				pl.Float64: "0.00",
				pl.Date: "mm/dd/yyyy",
				pl.Datetime: "mm/dd/yyyy hh:mm:ss",
				pl.Time: "hh:mm:ss",
			}
			call_logs_formats = {
				pl.Int8: "0",
				pl.Int16: "0",
				pl.Int32: "0",
				pl.Int64: "0",
				pl.Float32: "0",
				pl.Float64: "0",
				pl.Date: "mm/dd/yyyy",
				pl.Datetime: "mm/dd/yyyy hh:mm:ss",
			}
			st.session_state.step2_daily_remark_bytes = polars_to_excel_bytes(
				daily_remark,
				daily_remark_formats,
				left_align_col_indices=[5, 9],
			)
			st.session_state.step2_call_logs_bytes = polars_to_excel_bytes(
				call_logs,
				call_logs_formats,
				left_align_col_indices=[5, 9],
			)
			set_progress(40, "Extracting columns...", "Daily Remark and Daily Call Logs are ready.")
			max_date_value = daily_remark.select(pl.col("Date").max()).item() if "Date" in daily_remark.columns else None
			date_token = format_date_token(max_date_value)
			st.session_state.step2_daily_remark_name = f"maya_daily_remark_{date_token}.xlsx"
			st.session_state.step2_call_logs_name = f"maya_call_logs_{date_token}.xlsx"

			set_progress(46, "Matching data...", "Adding Call Type mapping to Call Logs.")

			step3_source_df = call_logs.to_pandas()
			step3_output_df = build_step3_call_type_file(step3_source_df)
			step3_source_file = BytesIO(st.session_state.step2_call_logs_bytes)
			step3_source_file.name = st.session_state.step2_call_logs_name or "maya_call_logs.xlsx"
			if Path(step3_source_file.name).suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
				st.session_state.step3_call_type_bytes = add_step3_call_type_preserve_workbook(step3_source_file, workbook_password=None)
			else:
				st.session_state.step3_call_type_bytes = to_output_excel_raw(step3_output_df, sheet_name="Call Logs")
			date_col = pick_column(step3_output_df.columns.tolist(), ["Date", "date", "DATE"])
			step3_max_date = None
			if date_col is not None:
				parsed_dates = pd.to_datetime(step3_output_df[date_col], errors="coerce")
				if parsed_dates.notna().any():
					step3_max_date = parsed_dates.max()
			st.session_state.step3_call_type_name = f"maya_call_logs_{format_date_token(step3_max_date)}.xlsx"

			set_progress(54, "Matching data...", "Call Type mapping complete.")
			set_progress(60, "Matching data...", "Getting columns from Endo file and merging accounts.")

			rewind_if_possible(selected_endo_source)
			agent_code_bcrm_volare, agent_code_volare_fullname = load_merge_account_mappings()
			merged_accounts = build_merged_accounts(
				endorsement_file=selected_endo_source,
				passkey=workbook_password,
				start_date=start_date,
				end_date=end_date,
				agent_code_bcrm_volare=agent_code_bcrm_volare,
				agent_code_volare_fullname=agent_code_volare_fullname,
			)
			step4_formatting = {
				"ACCOUNT NUMBER": "0",
				"OB": "0.00",
				"RECEIVED DATE": "mm/dd/yyyy",
				"AS_OF": "mm/dd/yyyy",
				"BIRTH_DATE": "mm/dd/yyyy",
				"PULLED OUT DATE": "mm/dd/yyyy",
			}
			st.session_state.step4_merged_accounts_bytes = save_xlsx(merged_accounts, step4_formatting)
			st.session_state.step4_merged_accounts_name = (
				f"maya_merged_accounts_{extract_mmddyy_token(selected_endo_name or '') or datetime.now().strftime('%m%d%y')}.xlsx"
			)

			set_progress(70, "Matching data...", "Merged accounts dataset is ready.")
			set_progress(76, "Combining datasets...", "Building the LOXON upload dataset.")

			status_ref, agent_ref = load_loxon_references()
			loxon_upload = build_loxon_upload(merged_accounts, daily_remark, status_ref, agent_ref)
			loxon_upload_formats = {
				pl.Int8: "0",
				pl.Int16: "0",
				pl.Int32: "0",
				pl.Int64: "0",
				pl.Float32: "0.00",
				pl.Float64: "0.00",
				pl.Date: "mm/dd/yyyy",
				pl.Datetime: "mm/dd/yyyy hh:mm:ss",
				pl.Time: "hh:mm:ss",
			}
			st.session_state.step5_loxon_upload_bytes = polars_to_excel_bytes(loxon_upload, loxon_upload_formats)
			max_event_dt = loxon_upload.select(pl.col("event_datetime_pht").max()).item() if "event_datetime_pht" in loxon_upload.columns else None
			if isinstance(max_event_dt, datetime):
				step5_date_token = max_event_dt.strftime("%Y-%m-%d")
			elif isinstance(max_event_dt, date):
				step5_date_token = max_event_dt.strftime("%Y-%m-%d")
			else:
				step5_date_token = datetime.now().strftime("%Y-%m-%d")
			st.session_state.step5_loxon_upload_name = f"MADRID_Feedback_{step5_date_token}.xlsx"

			set_progress(84, "Combining datasets...", "LOXON upload dataset complete.")
			set_progress(90, "Finalizing output...", "Filling missing values and preparing final file.")

			feedback_io = BytesIO(st.session_state.step5_loxon_upload_bytes)
			feedback_io.name = st.session_state.step5_loxon_upload_name or "MADRID_Feedback.xlsx"
			rewind_if_possible(selected_endo_source)
			final_bytes, final_unresolved_df, final_stats, final_name = fill_missing_values_pipeline(
				feedback_io,
				selected_endo_source,
				workbook_password=workbook_password,
			)

			st.session_state.final_output_bytes = final_bytes
			st.session_state.final_output_name = final_name
			st.session_state.final_unresolved_df = final_unresolved_df
			st.session_state.final_stats = final_stats

			set_progress(96, "Finalizing output...", "Final file prepared and stats generated.")
			set_progress(100, "Complete.", "Run finished.")
			st.success("Pipeline complete. Step downloads are ready below.")
		except Exception as exc:
			append_status(st.session_state.pipeline_percent, "Process failed.", f"{exc}")
			status_slot.markdown(f"**{st.session_state.pipeline_percent}%** Process failed.")
			st.error(f"Failed to process files: {exc}")
		finally:
			st.session_state.pipeline_running = False

	if st.session_state.step1_combined_daily_remark_bytes is not None:
		st.download_button(
			"Download Combined Daily Remark",
			data=st.session_state.step1_combined_daily_remark_bytes,
			file_name=st.session_state.step1_combined_daily_remark_name,
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	if st.session_state.step2_daily_remark_bytes is not None:
		st.download_button(
			"Download Extracted Daily Remark",
			data=st.session_state.step2_daily_remark_bytes,
			file_name=st.session_state.step2_daily_remark_name,
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	if st.session_state.step2_call_logs_bytes is not None:
		st.download_button(
			"Download Extracted Daily Call Logs",
			data=st.session_state.step2_call_logs_bytes,
			file_name=st.session_state.step2_call_logs_name,
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	if st.session_state.step3_call_type_bytes is not None:
		st.download_button(
			"Download Call Type Mapping",
			data=st.session_state.step3_call_type_bytes,
			file_name=st.session_state.step3_call_type_name,
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	if st.session_state.step4_merged_accounts_bytes is not None:
		st.download_button(
			"Download Merged Accounts",
			data=st.session_state.step4_merged_accounts_bytes,
			file_name=st.session_state.step4_merged_accounts_name,
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	if st.session_state.step5_loxon_upload_bytes is not None:
		st.download_button(
			"Download LOXON Upload",
			data=st.session_state.step5_loxon_upload_bytes,
			file_name=st.session_state.step5_loxon_upload_name,
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	if st.session_state.final_output_bytes is not None:
		st.download_button(
			"Download Final Output",
			data=st.session_state.final_output_bytes,
			file_name=st.session_state.final_output_name,
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			use_container_width=True,
		)

	if st.session_state.final_stats is not None:
		stats = st.session_state.final_stats
		st.success(
			f"Filled cpm_id: {stats['cpm_filled']:,} | Filled account_id: {stats['account_id_filled']:,} | "
			f"Filled number_contacted: {stats['contact_filled']:,} | Normalized number_contacted: {stats['normalized_contacts']:,}"
		)

	if st.session_state.final_unresolved_df is not None and not st.session_state.final_unresolved_df.empty:
		st.warning("Some rows could not be fully filled.")
		st.dataframe(st.session_state.final_unresolved_df, use_container_width=True)


run_ui()

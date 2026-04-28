from pathlib import Path
from io import BytesIO
from datetime import datetime, timedelta
import streamlit as st
import tempfile
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Directory helpers - FIXED: Use raw strings or double backslashes correctly
ACCOUNT_JOURNEY_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\ACCOUNT JOURNEY"
MERGE_ACCOUNTS_DIR = r"\\192.168.15.241\admin\ACTIVE\scperez\MAYA\MERGED ACCOUNTS\2026\APRIL"

def get_full_path(directory: str, filename: str) -> str:
    """Combine directory and filename into full path"""
    return str(Path(directory) / filename)

def get_yesterday_date_formatted():
    yesterday = datetime.now() - timedelta(days=1)
    return yesterday.strftime("%B %d, %Y").upper()

def get_today_date_formatted():
    today = datetime.now()
    return today.strftime("%B %d, %Y").upper()

# File helpers
def get_latest_file(directory, pattern):
    base = Path(directory)
    if not base.exists():
        return None

    files = list(base.glob(pattern))
    if not files:
        return None

    return max(files, key=lambda f: f.stat().st_mtime)

def get_latest_maj_file():
    return get_latest_file(
        ACCOUNT_JOURNEY_DIR,
        "maya_account_journey_*.xlsx"
    )

def get_latest_mma_file():
    latest = get_latest_file(
        MERGE_ACCOUNTS_DIR,
        "maya_merged_accounts_*.xlsx"
    )
    return latest.name if latest else ""

# ============================================================================
# FILE UTILITIES
# ============================================================================

def resolve_server_file(server_dir: str, server_input: str) -> tuple[Path | None, str | None]:
    """Resolve a server file by full path, relative path, or filename search."""
    base_dir = Path(server_dir)
    if not base_dir.exists() or not base_dir.is_dir():
        return None, f"Server folder is not reachable: {base_dir}"

    requested = (server_input or "").strip()
    if not requested:
        return None, "Please provide a server file name or path."

    normalized_requested = requested.replace("/", "\\")
    requested_path = Path(normalized_requested)
    
    if requested_path.exists() and requested_path.is_file():
        return requested_path, None
    
    candidate = base_dir / normalized_requested
    if candidate.exists() and candidate.is_file():
        return candidate, None
    
    requested_name = requested_path.name
    if not requested_name:
        return None, "Please provide a valid file name or relative path."
    
    matches = [p for p in base_dir.rglob("*") if p.is_file() and p.name.lower() == requested_name.lower()]
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        preview = "\n".join(str(p) for p in matches[:5])
        return None, f"Multiple files named '{requested_name}' were found. Please use a relative path.\nMatches:\n{preview}"
    
    return None, f"Server file not found from input: {requested}"
        
def read_excel_ws(file_path, sheet_name=None):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    data = list(ws.values)
    headers = data[0]
    rows = data[1:]

    return wb, ws, headers, rows

def append_rows_to_existing_excel(file_path, new_rows_df):
    """Append rows to existing Excel without changing any formatting."""
    wb = load_workbook(file_path)
    ws = wb.active

    # Append only the new rows (not the full dataframe)
    for row in dataframe_to_rows(new_rows_df, index=False, header=False):
        ws.append(row)

    wb.save(file_path)

def save_workbook(wb):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============================================================================
# UI SIDE
# ============================================================================

def automate_account_journey_update(maj_path, mma_path):
    try:
        import pandas as pd
        from openpyxl import load_workbook

        # ----------------------------
        # Resolve paths
        # ----------------------------
        if '\\' not in str(maj_path) and '/' not in str(maj_path):
            maj_path = Path(ACCOUNT_JOURNEY_DIR) / maj_path

        if '\\' not in str(mma_path) and '/' not in str(mma_path):
            mma_path = Path(MERGE_ACCOUNTS_DIR) / mma_path

        if not Path(maj_path).exists():
            return None, f"MAJ file not found: {maj_path}"

        if not Path(mma_path).exists():
            return None, f"MMA file not found: {mma_path}"

        # ----------------------------
        # Load with pandas (FAST)
        # ----------------------------
        maj_df = pd.read_excel(maj_path)
        mma_df = pd.read_excel(mma_path)

        maj_df.columns = maj_df.columns.astype(str).str.strip()
        mma_df.columns = mma_df.columns.astype(str).str.strip()

        # ----------------------------
        # Detect required columns
        # ----------------------------
        acc_col = next(c for c in mma_df.columns if "account" in c.lower())
        placement_col = next(c for c in mma_df.columns if "placement" in c.lower())
        ob_col = next(c for c in mma_df.columns if c.lower() in ["ob", "balance", "outstanding"])
        received_col = next(c for c in mma_df.columns if "received" in c.lower())
        fs_col = next(c for c in mma_df.columns if "fresh" in c.lower() or "spillover" in c.lower())

        # ----------------------------
        # FILTER ONLY FRESH
        # ----------------------------
        fresh_df = mma_df[mma_df[fs_col].astype(str).str.upper() == "FRESH"]

        st.write(f"Fresh records found: {len(fresh_df)}")

        if fresh_df.empty:
            return load_workbook(maj_path), "No FRESH records to add"

        # ----------------------------
        # BUILD FINAL STRUCTURE (MATCH MAJ FORMAT)
        # ----------------------------
        final_df = pd.DataFrame({
            "PRODUCT_NAME": fresh_df[placement_col],
            "Account #": fresh_df[acc_col],
            "Agent Name": None,
            "POUT Date": None,
            "Endo Date": fresh_df[received_col],
            "Balance": fresh_df[ob_col],
        })

        # ----------------------------
        # LOAD MAJ WORKBOOK (FOR SAFE APPEND)
        # ----------------------------
        wb = load_workbook(maj_path)
        ws = wb.active

        # ----------------------------
        # APPEND ONLY NEW ROWS (FORMAT SAFE)
        # ----------------------------
        # GET MAJ HEADER STRUCTURE (SOURCE OF TRUTH)
        headers = [str(c.value).strip() for c in ws[1]]

        # BUILD ROWS BASED ON MAJ STRUCTURE (NOT MMA ORDER)
        for _, row in fresh_df.iterrows():

            new_row = []

            for col in headers:

                val = None

                if col == "PRODUCT_NAME":
                    val = row.get(placement_col)

                elif col == "Account #":
                    val = row.get(acc_col)

                elif col == "Endo Date":
                    val = row.get(received_col)

                elif col == "Balance":
                    val = row.get(ob_col)

                elif col == "Agent Name":
                    val = None

                # IMPORTANT: ignore extra MMA columns (prevents shifting)
                else:
                    val = None

                new_row.append(val)

            ws.append(new_row)


        return wb, f"Updated successfully: {len(final_df)} rows added"

    except Exception as e:
        import traceback
        return None, f"Error: {str(e)}\n{traceback.format_exc()}"


# STREAMLIT UI - Only Account Journey Update
st.set_page_config(page_title="Account Journey Automation", layout="wide")
st.title("📊 Automation for ACCJOURNEY & Conanduncon")
st.caption("Complete Account Journey Update")

if "processed_data" not in st.session_state:
    st.session_state.processed_data = None


# Tabs: Main (MAJ/MMA) and Additional Files (MDR/CONANDUNCON)

tab1, tab2 = st.tabs(["Main Update", "Additional Files (Monthly)"])

with tab1:
    st.markdown("### Main Account Journey Update")

    with st.form("main_update_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            # Use yesterday's date instead of latest file
            default_maj_filename = f"maya_account_journey_{get_yesterday_date_formatted()}.xlsx"
            maj_path = st.text_input(
                "MAJ File (maya_account_journey_*.xlsx)",
                value=default_maj_filename,  # ← NOW USING YESTERDAY'S DATE
                help=f"Full path: {ACCOUNT_JOURNEY_DIR}\n\nAuto-populated with yesterday's date: {get_yesterday_date_formatted()}"
            )
            uploaded_maj = st.file_uploader("Or upload MAJ file", type=["xlsx"], key="main_upload_maj")
        
        with col2:
            # Keep MMA as latest or change to yesterday as well?
            default_mma_filename = f"maya_merged_accounts_{datetime.now().strftime('%m%d%y')}.xlsx"
            mma_path = st.text_input(
                "MMA File (maya_merged_accounts_*.xlsx)",
                value=get_latest_mma_file(),  # Keep as is or change to default_mma_filename
                help=f"Full path: {MERGE_ACCOUNTS_DIR}\n\nAuto-suggests the latest file in the Merge Accounts directory."
            )
            uploaded_mma = st.file_uploader("Or upload MMA file", type=["xlsx"], key="main_upload_mma")
        
        submitted_main = st.form_submit_button("Submit", use_container_width=True, type="primary")

    if submitted_main:
        # Determine sources - check if uploaded or text input
        maj_source = uploaded_maj if uploaded_maj else maj_path
        mma_source = uploaded_mma if uploaded_mma else mma_path
        
        if not maj_source or not mma_source:
            st.error("Please provide both Maya Account Journey and Maya Merged Accounts files.")
        else:
            try:
                # Handle uploaded files
                temp_files = []
                
                if uploaded_maj:
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                    tmp.write(uploaded_maj.getvalue())
                    tmp.close()
                    maj_source = tmp.name
                    temp_files.append(tmp.name)
                else:
                    # For text input (filename only), combine with directory path
                    # Check if it's just a filename (no backslashes or forward slashes)
                    if '\\' not in str(maj_source) and '/' not in str(maj_source):
                        maj_source = str(Path(ACCOUNT_JOURNEY_DIR) / maj_source)
                        st.info(f"Looking for MAJ file at: {maj_source}")
                
                if uploaded_mma:
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                    tmp.write(uploaded_mma.getvalue())
                    tmp.close()
                    mma_source = tmp.name
                    temp_files.append(tmp.name)
                else:
                    # For text input (filename only), combine with directory path
                    if '\\' not in str(mma_source) and '/' not in str(mma_source):
                        mma_source = str(Path(MERGE_ACCOUNTS_DIR) / mma_source)
                        st.info(f"Looking for MMA file at: {mma_source}")
                
                # Check if files exist before proceeding
                if not uploaded_maj and not Path(maj_source).exists():
                    st.error(f"❌ MAJ file not found: {maj_source}\n\nPlease check that the file exists in the directory.")
                    st.stop()
                
                if not uploaded_mma and not Path(mma_source).exists():
                    st.error(f"❌ MMA file not found: {mma_source}\n\nPlease check that the file exists in the directory.")
                    st.stop()
                
                with st.spinner("Running Main Account Journey Update..."):
                    updated_maj, result_message = automate_account_journey_update(
                        maj_source, mma_source
                    )
                    
                    if updated_maj is not None:
                        st.success(result_message)
                        
                        output = save_workbook(updated_maj)

                        filename = f"maya_account_journey_{get_today_date_formatted()}.xlsx"

                        st.download_button(
                            "📥 Download Updated MAJ",
                            data=output.getvalue(),
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                        st.caption(f"📅 Downloaded file: {filename}")

                    else:
                        st.error(result_message)
                
                # Cleanup temp files
                for temp_file in temp_files:
                    try:
                        Path(temp_file).unlink(missing_ok=True)
                    except:
                        pass
                        
            except Exception as exc:
                st.error(f"❌ Automation failed: {exc}")
                st.write("Debug info - maj_source:", maj_source)
                st.write("Debug info - mma_source:", mma_source)

with tab2:
    st.markdown("### Additional Files (Monthly Use)")
    st.info("""
    **This tab is for monthly updates only** (typically on the 1st of each month).
    
    Additional files used:
    - **MDR** (maya_daily_remark.xlsx) - Daily remarks and notes
    - **CONANDUNCON** - Consolidated and unconsolidated data
    
    These files supplement the main update with additional context and data.
    """)
    
    with st.form("additional_files_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            mdr_filename = st.text_input(
                "MDR File (maya_daily_remark.xlsx)",
                value="maya_daily_remark.xlsx",
                help=f"Full path: {ACCOUNT_JOURNEY_DIR}\n\nEnter just the filename (e.g., maya_daily_remark.xlsx)"
            )
            uploaded_mdr = st.file_uploader("Or upload MDR file", type=["xlsx"], key="upload_mdr")
        
        with col2:
            conanduncon_filename = st.text_input(
                "CONANDUNCON File",
                value="CONANDUNCON_40222026.xlsx",
                help=f"Full path: {ACCOUNT_JOURNEY_DIR}\n\nEnter just the filename (e.g., CONANDUNCON_40222026.xlsx)"
            )
            uploaded_conanduncon = st.file_uploader("Or upload CONANDUNCON file", type=["xlsx"], key="upload_conanduncon")
        
        st.warning("⚠️ Note: This tab saves file paths for reference. Use the Main Update tab to run the actual update.")
        
        submitted_additional = st.form_submit_button("💾 Save Additional File Paths", use_container_width=True, type="primary")
    
    if submitted_additional:
        # FIX: Use correct variable names (mdr_filename / conanduncon_filename, not mdr_path / conanduncon_path)
        mdr_path = get_full_path(ACCOUNT_JOURNEY_DIR, mdr_filename) if mdr_filename else None
        conanduncon_path = get_full_path(ACCOUNT_JOURNEY_DIR, conanduncon_filename) if conanduncon_filename else None

        st.session_state.mdr_path = mdr_path
        st.session_state.conanduncon_path = conanduncon_path

        if uploaded_mdr:
            st.success(f"MDR file '{uploaded_mdr.name}' ready for processing")
        if uploaded_conanduncon:
            st.success(f"CONANDUNCON file '{uploaded_conanduncon.name}' ready for processing")

        if not uploaded_mdr and not uploaded_conanduncon:
            st.success("Additional file paths saved successfully!")

        if uploaded_mdr:
            try:
                mdr_df = pd.read_excel(uploaded_mdr)
                with st.expander("Preview MDR File"):
                    st.dataframe(mdr_df.head(10), use_container_width=True)
            except Exception as e:
                st.error(f"Error previewing MDR: {e}")

        if uploaded_conanduncon:
            try:
                con_df = pd.read_excel(uploaded_conanduncon)
                with st.expander("Preview CONANDUNCON File"):
                    st.dataframe(con_df.head(10), use_container_width=True)
            except Exception as e:
                st.error(f"Error previewing CONANDUNCON: {e}")

st.markdown("---")
st.caption("Account Journey Automation Tool | Lyra Format Compliant | v2.1")